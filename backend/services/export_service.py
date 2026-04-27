"""
Service pour l'export comptable.
Genere des archives ZIP contenant operations (CSV/PDF/Excel),
releves bancaires et justificatifs pour un ou plusieurs mois.

Regles comptables :
- Les ops "perso" sont exclues des totaux BNC (section separee)
- Les ops sans categorie / "Autres" vont en compte d'attente
- Les ventilations sont explosees en sous-lignes
"""
from __future__ import annotations

import io
import json
import logging
import os
import re
import shutil
import zipfile
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Optional, List

from backend.core.config import (
    IMPORTS_RELEVES_DIR, EXPORTS_DIR, RAPPORTS_DIR, REPORTS_DIR,
    JUSTIFICATIFS_TRAITES_DIR, MOIS_FR, ASSETS_DIR, APP_NAME,
    ensure_directories,
)
from backend.services import operation_service, report_service

logger = logging.getLogger(__name__)


# ─── Helpers utilitaires ───

def _format_amount_fr(amount: float) -> str:
    """1234.56 -> '1 234,56', 0 -> '0,00'"""
    if amount == 0:
        return "0,00"
    formatted = f"{amount:,.2f}"  # "1,234.56"
    formatted = formatted.replace(",", " ").replace(".", ",")
    return formatted


def _export_filename(year: int, month: int, ext: str) -> str:
    """-> 'Export_Comptable_2025_Janvier.csv'"""
    mois_label = MOIS_FR[month - 1].capitalize() if 1 <= month <= 12 else f"{month:02d}"
    return f"Export_Comptable_{year}_{mois_label}.{ext}"


def _get_justificatif_name(op: dict) -> str:
    """Extrait le basename du justificatif ou retourne vide. Marque [R] les reconstitués."""
    lien = op.get("Lien justificatif", "") or op.get("justificatif", "") or ""
    if not lien:
        return ""
    basename = os.path.basename(lien)
    # Détection fac-similé : nouveau format `_fs` ou legacy `reconstitue_`
    try:
        from backend.services import rename_service
        if rename_service.is_facsimile(basename):
            return f"{basename} [R]"
    except Exception:
        if basename.startswith("reconstitue_") or "_fs.pdf" in basename.lower():
            return f"{basename} [R]"
    return basename


def _safe_float(val) -> float:
    """Convertit une valeur en float en gerant NaN et None."""
    if val is None:
        return 0.0
    try:
        f = float(val)
        if f != f:  # NaN check
            return 0.0
        return f
    except (ValueError, TypeError):
        return 0.0


# ─── Preparation des operations ───

def _prepare_export_operations(operations: list, filename: str) -> dict:
    """
    Prepare les operations pour l'export en les classant en 3 groupes.
    Explose les ventilations. Trie par date ASC.

    Returns:
        {
            "pro": [...],
            "perso": [...],
            "attente": [...],
            "totals": { recettes_pro, charges_pro, solde_bnc, total_perso, nb_perso, total_attente, nb_attente }
        }
    """
    pro = []
    perso = []
    attente = []

    for op in operations:
        ventilation = op.get("ventilation") or []

        if ventilation:
            # Exploser les sous-lignes
            n = len(ventilation)
            for i, vline in enumerate(ventilation):
                sub = {
                    "Date": op.get("Date", ""),
                    "Libelle": f"{op.get('Libellé', op.get('Libelle', ''))} [V{i+1}/{n}]",
                    "Debit": _safe_float(vline.get("montant", 0)) if _safe_float(op.get("Débit", 0)) > 0 else 0.0,
                    "Credit": _safe_float(vline.get("montant", 0)) if _safe_float(op.get("Crédit", 0)) > 0 else 0.0,
                    "Categorie": vline.get("categorie", ""),
                    "Sous_categorie": vline.get("sous_categorie", ""),
                    "Justificatif": os.path.basename(vline.get("justificatif", "") or ""),
                    "Commentaire": op.get("Commentaire", ""),
                }
                _classify_line(sub, pro, perso, attente)
        else:
            line = {
                "Date": op.get("Date", ""),
                "Libelle": op.get("Libellé", op.get("Libelle", "")),
                "Debit": _safe_float(op.get("Débit", 0)),
                "Credit": _safe_float(op.get("Crédit", 0)),
                "Categorie": op.get("Catégorie", op.get("Categorie", "")),
                "Sous_categorie": op.get("Sous-catégorie", op.get("Sous_categorie", "")),
                "Justificatif": _get_justificatif_name(op),
                "Commentaire": op.get("Commentaire", ""),
            }
            _classify_line(line, pro, perso, attente)

    # Trier chaque groupe par date ASC
    for group in [pro, perso, attente]:
        group.sort(key=lambda x: x.get("Date", ""))

    # Calculer les totaux
    recettes_pro = sum(l["Credit"] for l in pro)
    charges_pro = sum(l["Debit"] - (l.get("csg_non_deductible") or 0) for l in pro)

    totals = {
        "recettes_pro": recettes_pro,
        "charges_pro": charges_pro,
        "solde_bnc": recettes_pro - charges_pro,
        "total_perso": sum(l["Debit"] + l["Credit"] for l in perso),
        "nb_perso": len(perso),
        "total_attente": sum(l["Debit"] + l["Credit"] for l in attente),
        "nb_attente": len(attente),
        "debit_perso": sum(l["Debit"] for l in perso),
        "credit_perso": sum(l["Credit"] for l in perso),
        "debit_attente": sum(l["Debit"] for l in attente),
        "credit_attente": sum(l["Credit"] for l in attente),
    }

    return {"pro": pro, "perso": perso, "attente": attente, "totals": totals}


def get_bnc_summary(year: int, ca_liasse: Optional[float] = None) -> dict:
    """Façade vers `bnc_service.compute_bnc(year)` pour les consommateurs externes.

    Retourne le breakdown sérialisé en dict (cohérent avec les valeurs exposées par
    le dashboard via `get_dashboard_data(..., year_full=year)`).
    """
    from backend.services import bnc_service
    breakdown = bnc_service.compute_bnc(year, ca_liasse=ca_liasse)
    return breakdown.to_dict()


def _classify_line(line: dict, pro: list, perso: list, attente: list):
    """Classe une ligne normalisee dans le bon groupe.

    Les catégories `EXCLUDED_FROM_CHARGES_PRO` (Immobilisations, Dotations aux
    amortissements, Ventilé) sont placées en 'attente' — visibles dans le compte
    d'attente de l'export comptable mais hors agrégat `charges_pro` du BNC.
    Évite le double-comptage avec les dotations annuelles calculées par le module
    Amortissements.
    """
    # Import local pour éviter un cycle
    from backend.services.analytics_service import EXCLUDED_FROM_CHARGES_PRO

    cat = (line.get("Categorie") or "").strip()
    cat_lower = cat.lower()

    if cat_lower == "perso":
        perso.append(line)
    elif cat == "" or cat_lower == "autres" or cat in EXCLUDED_FROM_CHARGES_PRO:
        attente.append(line)
    else:
        pro.append(line)


# ─── Titre auto ───

MONTH_NAMES_FR = [
    'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
    'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre',
]


def build_export_title(year: int, month: int) -> str:
    """Construit le titre auto pour un export mensuel."""
    return f"Toutes catégories — {MONTH_NAMES_FR[month - 1]} {year}"


# ─── Historique des exports ───

EXPORTS_HISTORY_FILE = EXPORTS_DIR / "exports_history.json"


def _load_exports_history() -> list:
    """Charge l'historique des exports."""
    if not EXPORTS_HISTORY_FILE.exists():
        return []
    try:
        data = json.loads(EXPORTS_HISTORY_FILE.read_text(encoding="utf-8"))
        return data.get("exports", [])
    except Exception:
        return []


def _save_exports_history(exports: list) -> None:
    """Sauvegarde l'historique des exports."""
    ensure_directories()
    EXPORTS_HISTORY_FILE.write_text(
        json.dumps({"exports": exports}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _log_export(
    year: int,
    month: int,
    fmt: str,
    filename: str,
    title: str,
    nb_operations: int,
) -> None:
    """Ajoute une entree dans l'historique des exports.

    Déduplique par (year, month, format) : toute entrée antérieure pour le même
    triplet est retirée et le ZIP correspondant est supprimé du disque (si encore
    présent dans EXPORTS_DIR). Évite l'empilement des exports successifs pour un
    même mois.
    """
    history = _load_exports_history()

    # Dédup : supprimer les entrées antérieures du même (year, month, fmt)
    retained = []
    superseded = []
    for e in history:
        if (
            e.get("year") == year
            and e.get("month") == month
            and e.get("format") == fmt
        ):
            superseded.append(e)
        else:
            retained.append(e)

    # Supprimer les ZIP des entrées supplantées (best-effort)
    for old in superseded:
        old_fn = old.get("filename")
        if not old_fn:
            continue
        old_path = EXPORTS_DIR / old_fn
        try:
            if old_path.exists() and old_path.suffix == ".zip":
                old_path.unlink()
        except OSError as err:
            logger.warning("Impossible de supprimer l'export supplanté %s: %s", old_fn, err)

    retained.append({
        "id": f"exp_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
        "year": year,
        "month": month,
        "format": fmt,
        "filename": filename,
        "title": title,
        "nb_operations": nb_operations,
        "generated_at": datetime.now().isoformat(),
    })
    _save_exports_history(retained)


def get_exports_history(year: Optional[int] = None) -> list:
    """Retourne l'historique des exports, filtre optionnel par annee."""
    history = _load_exports_history()
    if year is not None:
        history = [e for e in history if e.get("year") == year]
    return history


def get_month_export_status(year: int) -> dict:
    """
    Pour chaque mois 1-12, retourne le statut des exports avec preview du contenu.
    Croise les fichiers d'operations avec l'historique.
    """
    ensure_directories()

    # Compter les ops par mois et garder les fichiers
    op_files = operation_service.list_operation_files()
    ops_by_month: dict = {}
    files_by_month: dict = {}
    for f in op_files:
        if f.get("year") == year:
            m = f.get("month")
            if m:
                ops_by_month[m] = f.get("count", 0)
                files_by_month[m] = f

    # Charger l'historique
    history = _load_exports_history()
    year_history = [e for e in history if e.get("year") == year]

    months = []
    for m in range(1, 13):
        nb_ops = ops_by_month.get(m, 0)
        has_data = m in ops_by_month

        # Trouver les derniers exports PDF et CSV
        pdf_entries = [e for e in year_history if e["month"] == m and e["format"] == "pdf"]
        csv_entries = [e for e in year_history if e["month"] == m and e["format"] == "csv"]

        last_pdf = pdf_entries[-1] if pdf_entries else None
        last_csv = csv_entries[-1] if csv_entries else None

        # Verifier que les fichiers existent encore
        has_pdf = bool(last_pdf and (EXPORTS_DIR / last_pdf["filename"]).exists())
        has_csv = bool(last_csv and (EXPORTS_DIR / last_csv["filename"]).exists())

        # Preview du contenu ZIP
        nb_releves = 0
        nb_rapports = 0
        nb_justificatifs = 0
        if has_data:
            target = files_by_month.get(m)
            if target:
                bank_pdf = _find_bank_statement(target["filename"])
                nb_releves = 1 if bank_pdf and bank_pdf.exists() else 0
                month_name = MONTH_NAMES_FR[m - 1]
                nb_rapports = len(_find_existing_reports(year, m, month_name))
                try:
                    operations = operation_service.load_operations(target["filename"])
                    nb_justificatifs = len(_collect_justificatifs(operations))
                except Exception:
                    pass

        months.append({
            "month": m,
            "label": MONTH_NAMES_FR[m - 1],
            "nb_operations": nb_ops,
            "has_data": has_data,
            "has_pdf": has_pdf,
            "has_csv": has_csv,
            "last_pdf_filename": last_pdf["filename"] if has_pdf else None,
            "last_pdf_date": last_pdf["generated_at"] if has_pdf else None,
            "last_csv_filename": last_csv["filename"] if has_csv else None,
            "last_csv_date": last_csv["generated_at"] if has_csv else None,
            "nb_releves": nb_releves,
            "nb_rapports": nb_rapports,
            "nb_justificatifs": nb_justificatifs,
        })

    return {"year": year, "months": months}


def _collect_justificatifs(operations: list) -> list:
    """Collecte les chemins des justificatifs existants pour des operations."""
    found = []
    seen = set()
    for op in operations:
        lien = op.get("Lien justificatif", "")
        if lien:
            just_filename = Path(lien).name
            if just_filename in seen:
                continue
            seen.add(just_filename)
            just_path = JUSTIFICATIFS_TRAITES_DIR / just_filename
            if just_path.exists():
                found.append(just_path)
    return found


def get_export_preview(year: int, month: int) -> dict:
    """
    Retourne le contenu qui sera inclus dans le ZIP pour un mois donne.
    """
    ensure_directories()

    op_files = operation_service.list_operation_files()
    target_file = None
    for f in op_files:
        if f.get("year") == year and f.get("month") == month:
            target_file = f
            break

    if not target_file:
        return {"nb_operations": 0, "nb_releves": 0, "nb_rapports": 0, "nb_justificatifs": 0}

    operations = operation_service.load_operations(target_file["filename"])

    # Releve bancaire
    bank_pdf = _find_bank_statement(target_file["filename"])
    nb_releves = 1 if bank_pdf and bank_pdf.exists() else 0

    # Rapports
    month_name = MONTH_NAMES_FR[month - 1]
    nb_rapports = len(_find_existing_reports(year, month, month_name))

    # Justificatifs
    justificatifs = _collect_justificatifs(operations)

    return {
        "nb_operations": len(operations),
        "nb_releves": nb_releves,
        "nb_rapports": nb_rapports,
        "nb_justificatifs": len(justificatifs),
    }


def _find_existing_reports(year: int, month: int, month_name: str) -> list:
    """Trouve les rapports existants pour un mois. Retourne une liste de Paths."""
    found = []
    month_name_lower = month_name.lower()
    # Exclure les fichiers déjà inclus séparément dans le ZIP
    excluded_prefixes = ("export_comptable_", "compte_attente_")
    for reports_dir in [RAPPORTS_DIR, REPORTS_DIR]:
        if not reports_dir.exists():
            continue
        for f in reports_dir.iterdir():
            if f.suffix in (".pdf", ".csv", ".xlsx"):
                name_lower = f.name.lower()
                if name_lower.startswith(excluded_prefixes):
                    continue
                if (month_name_lower in name_lower or
                    f"_{year}_{month:02d}_" in name_lower or
                    f"_{month:02d}_{year}" in name_lower):
                    found.append(f)
    return found


def get_available_reports_for_month(year: int, month: int) -> dict:
    """Retourne les rapports disponibles pour inclusion dans un export mensuel."""
    month_name = MONTH_NAMES_FR[month - 1]
    auto_detected = _find_existing_reports(year, month, month_name)
    auto_filenames = {f.name for f in auto_detected}

    all_reports = report_service.get_all_reports()

    results = []
    seen: set = set()

    # Auto-detected first
    for f in auto_detected:
        # Enrich with metadata from index if available
        meta = next((r for r in all_reports if r.get("filename") == f.name), None)
        results.append({
            "filename": f.name,
            "title": meta.get("title", f.stem) if meta else f.stem,
            "auto_detected": True,
            "format": f.suffix.lstrip("."),
            "year": meta.get("year") if meta else year,
            "month": meta.get("month") if meta else month,
        })
        seen.add(f.name)

    # Gallery reports not auto-detected
    for r in all_reports:
        fn = r.get("filename", "")
        if fn and fn not in seen:
            results.append({
                "filename": fn,
                "title": r.get("title", fn),
                "auto_detected": False,
                "format": r.get("format", ""),
                "year": r.get("year"),
                "month": r.get("month"),
            })

    return {"year": year, "month": month, "reports": results}


def generate_single_export(
    year: int,
    month: int,
    fmt: str,
    report_filenames: Optional[List[str]] = None,
    include_compte_attente: bool = True,
) -> dict:
    """
    Genere un export ZIP pour un mois donne.
    Architecture : racine = CSV/PDF, dossiers releves/, rapports/, justificatifs/.
    report_filenames: None = auto-discovery, [] = no reports, [...] = explicit list.
    """
    ensure_directories()

    op_files = operation_service.list_operation_files()
    target_file = None
    for f in op_files:
        if f.get("year") == year and f.get("month") == month:
            target_file = f
            break

    if not target_file:
        raise ValueError(f"Aucune opération trouvée pour {month:02d}/{year}")

    operations = operation_service.load_operations(target_file["filename"])
    if not operations:
        raise ValueError(f"Fichier vide pour {month:02d}/{year}")

    month_name = MONTH_NAMES_FR[month - 1]
    title = build_export_title(year, month)
    prepared = _prepare_export_operations(operations, target_file["filename"])
    nb_ops = len(operations)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    zip_name = f"Export_Comptable_{year}_{month_name}_{fmt.upper()}_{timestamp}.zip"
    zip_path = EXPORTS_DIR / zip_name

    files_included = []

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        # ── Operations PDF + CSV (toujours les deux) ──
        pdf_bytes = _generate_pdf_content(prepared, month_name, year, month)
        pdf_arc = _export_filename(year, month, "pdf")
        zf.writestr(pdf_arc, pdf_bytes)
        files_included.append({"name": pdf_arc, "type": "pdf"})

        csv_content = _generate_csv_content(prepared, month_name, year)
        csv_arc = _export_filename(year, month, "csv")
        zf.writestr(csv_arc, csv_content)
        files_included.append({"name": csv_arc, "type": "csv"})

        # ── Ventilation par catégorie (PDF + CSV avec sous-totaux) ──
        try:
            cat_pdf_bytes = _generate_pdf_by_category(prepared, month_name, year, month)
            cat_pdf_arc = _category_pdf_filename(year, month)
            zf.writestr(cat_pdf_arc, cat_pdf_bytes)
            files_included.append({"name": cat_pdf_arc, "type": "ventilation_categorie_pdf"})

            cat_csv_content = _generate_csv_by_category(prepared, month_name, year)
            cat_csv_arc = _category_csv_filename(year, month)
            zf.writestr(cat_csv_arc, cat_csv_content)
            files_included.append({"name": cat_csv_arc, "type": "ventilation_categorie_csv"})
        except Exception as e:
            logger.warning("Impossible de générer la ventilation par catégorie: %s", e)

        # ── Copie standalone dans REPORTS_DIR + enregistrement GED ──
        try:
            from backend.services import ged_service
            from backend.core.config import REPORTS_DIR as _REPORTS_DIR
            _REPORTS_DIR.mkdir(parents=True, exist_ok=True)

            for arc_name_r, content_r, fmt_r in [
                (pdf_arc, pdf_bytes, "pdf"),
                (csv_arc, csv_content.encode("utf-8") if isinstance(csv_content, str) else csv_content, "csv"),
            ]:
                standalone_path = _REPORTS_DIR / arc_name_r
                if isinstance(content_r, str):
                    standalone_path.write_text(content_r, encoding="utf-8")
                else:
                    standalone_path.write_bytes(content_r)

                # Deduplication GED : chercher un export comptable existant pour ce mois/format
                existing_meta = ged_service.load_metadata()
                existing_docs = existing_meta.get("documents", {})
                replaced_fn = None
                for _did, _doc in existing_docs.items():
                    if _doc.get("type") != "rapport":
                        continue
                    _rm = _doc.get("rapport_meta") or {}
                    _fl = _rm.get("filters") or {}
                    if (
                        _fl.get("report_type") == "export_comptable"
                        and _fl.get("year") == year
                        and _fl.get("month") == month
                        and _rm.get("format") == fmt_r
                    ):
                        replaced_fn = _doc.get("filename") or Path(_did).name
                        break

                ged_service.register_rapport(
                    filename=arc_name_r,
                    path=str(standalone_path),
                    title=f"Export Comptable — {month_name} {year}",
                    description=f"{nb_ops} opérations",
                    filters={
                        "year": year,
                        "month": month,
                        "report_type": "export_comptable",
                        "categories": ["__all__"],
                    },
                    format_type=fmt_r,
                    template_id=None,
                    replaced_filename=replaced_fn,
                )
        except Exception as e:
            logger.warning("Impossible d'enregistrer l'export comptable dans la GED: %s", e)

        # ── Releves bancaires ──
        bank_pdf = _find_bank_statement(target_file["filename"])
        if bank_pdf and bank_pdf.exists():
            arc_name = f"releves/{bank_pdf.name}"
            zf.write(str(bank_pdf), arcname=arc_name)
            files_included.append({"name": arc_name, "type": "releve"})

        # ── Rapports ──
        if report_filenames is None:
            reports = _find_existing_reports(year, month, month_name)
        else:
            reports = []
            for fn in report_filenames:
                rp = report_service.get_report_path(fn)
                if rp and rp.exists():
                    reports.append(rp)
                else:
                    logger.warning("Report file not found: %s", fn)
        for rp in reports:
            arc_name = f"rapports/{rp.name}"
            zf.write(str(rp), arcname=arc_name)
            files_included.append({"name": arc_name, "type": "rapport"})

        # ── Justificatifs ──
        justificatifs = _collect_justificatifs(operations)
        for jp in justificatifs:
            arc_name = f"justificatifs/{jp.name}"
            zf.write(str(jp), arcname=arc_name)
        if justificatifs:
            files_included.append({"name": f"justificatifs/ ({len(justificatifs)} fichiers)", "type": "justificatifs"})

        # ── Compte d'attente ──
        if include_compte_attente:
            try:
                from backend.services import alerte_export_service
                for ca_fmt in ("pdf", "csv"):
                    ca_result = alerte_export_service.export_compte_attente(year, month, ca_fmt)
                    ca_filename = ca_result["filename"]
                    ca_path = EXPORTS_DIR / ca_filename
                    if ca_path.exists():
                        arc_name = f"compte_attente/{ca_filename}"
                        zf.write(str(ca_path), arcname=arc_name)
                        files_included.append({"name": arc_name, "type": "compte_attente"})
            except Exception as e:
                logger.warning("Impossible de generer le compte d'attente: %s", e)

    _log_export(year, month, fmt, zip_name, title, nb_ops)

    return {
        "filename": zip_name,
        "title": title,
        "nb_operations": nb_ops,
        "generated": True,
        "download_url": f"/api/exports/download/{zip_name}",
        "size_human": _format_size(zip_path.stat().st_size),
        "files_included": files_included,
    }


def generate_batch_export(year: int, months: list, fmt: str) -> dict:
    """
    Genere un lot d'exports pour plusieurs mois dans un seul ZIP.
    Architecture: {Mois}/operations.{fmt}, {Mois}/releves/, {Mois}/rapports/, {Mois}/justificatifs/
    """
    ensure_directories()

    generated_count = 0
    already_existed = 0
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    zip_name = f"Exports_Comptable_{year}_{timestamp}.zip"
    zip_path = EXPORTS_DIR / zip_name

    op_files = operation_service.list_operation_files()

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for m in months:
            month_name = MONTH_NAMES_FR[m - 1]
            prefix = f"{month_name}_{year}"

            target_file = None
            for f in op_files:
                if f.get("year") == year and f.get("month") == m:
                    target_file = f
                    break
            if not target_file:
                continue

            try:
                operations = operation_service.load_operations(target_file["filename"])
                if not operations:
                    continue
                prepared = _prepare_export_operations(operations, target_file["filename"])

                # Operations
                if fmt == "pdf":
                    pdf_bytes = _generate_pdf_content(prepared, month_name, year, m)
                    zf.writestr(f"{prefix}/{_export_filename(year, m, 'pdf')}", pdf_bytes)
                else:
                    csv_content = _generate_csv_content(prepared, month_name, year)
                    zf.writestr(f"{prefix}/{_export_filename(year, m, 'csv')}", csv_content)

                # Ventilation par catégorie (toujours PDF + CSV)
                try:
                    zf.writestr(
                        f"{prefix}/{_category_pdf_filename(year, m)}",
                        _generate_pdf_by_category(prepared, month_name, year, m),
                    )
                    zf.writestr(
                        f"{prefix}/{_category_csv_filename(year, m)}",
                        _generate_csv_by_category(prepared, month_name, year),
                    )
                except Exception as e:
                    logger.warning("Batch ventilation par catégorie skip %d: %s", m, e)

                # Releves
                bank_pdf = _find_bank_statement(target_file["filename"])
                if bank_pdf and bank_pdf.exists():
                    zf.write(str(bank_pdf), arcname=f"{prefix}/releves/{bank_pdf.name}")

                # Rapports
                reports = _find_existing_reports(year, m, month_name)
                for rp in reports:
                    zf.write(str(rp), arcname=f"{prefix}/rapports/{rp.name}")

                # Justificatifs
                justificatifs = _collect_justificatifs(operations)
                for jp in justificatifs:
                    zf.write(str(jp), arcname=f"{prefix}/justificatifs/{jp.name}")

                generated_count += 1
                title = build_export_title(year, m)
                _log_export(year, m, fmt, zip_name, title, len(operations))

            except Exception as e:
                logger.warning("Batch export skip month %d: %s", m, e)

    return {
        "zip_filename": zip_name,
        "generated_count": generated_count,
        "already_existed": already_existed,
        "total": len(months),
        "download_url": f"/api/exports/download/{zip_name}",
    }


# ─── Periodes disponibles ───

def get_available_periods() -> dict:
    """
    Retourne les mois disponibles et leur statut.
    """
    ensure_directories()

    op_files = operation_service.list_operation_files()
    periods = []

    for f in op_files:
        month = f.get("month")
        year = f.get("year")
        if not month or not year:
            continue

        try:
            ops = operation_service.load_operations(f["filename"])
            total = len(ops)
            with_just = sum(1 for op in ops if op.get("Justificatif"))
            ratio = (with_just / total * 100) if total > 0 else 0.0
        except Exception:
            ratio = 0.0

        month_name = MOIS_FR[month - 1].capitalize() if 1 <= month <= 12 else str(month)

        periods.append({
            "year": year,
            "month": month,
            "month_name": month_name,
            "filename": f["filename"],
            "count": f.get("count", 0),
            "total_debit": f.get("total_debit", 0),
            "total_credit": f.get("total_credit", 0),
            "has_export": _has_export(year, month),
            "justificatif_ratio": round(ratio, 1),
        })

    periods.sort(key=lambda p: (p["year"], p["month"]), reverse=True)
    years = sorted(set(p["year"] for p in periods), reverse=True)

    return {"periods": periods, "years": years}


def _has_export(year: int, month: int) -> bool:
    """Verifie si un export existe deja pour ce mois."""
    if not EXPORTS_DIR.exists():
        return False
    # Nouveau format
    pattern_new = f"Export_Comptable_{year}_{MOIS_FR[month - 1].capitalize() if 1 <= month <= 12 else f'{month:02d}'}*"
    # Ancien format (retrocompatibilite)
    pattern_old = f"export_{year}_{month:02d}_*"
    return (
        len(list(EXPORTS_DIR.glob(pattern_new))) > 0
        or len(list(EXPORTS_DIR.glob(pattern_old))) > 0
    )


# ─── Listing exports existants ───

def _build_releve_display_map() -> dict:
    """Construit un mapping hash → 'Relevé Mois Année' depuis les fichiers d'opérations."""
    mapping: dict = {}
    try:
        op_files = operation_service.list_operation_files()
        for f in op_files:
            m = re.search(r"_([a-f0-9]{8})\.json$", f.get("filename", ""))
            if m:
                file_hash = m.group(1)
                year = f.get("year")
                month = f.get("month")
                if year and month and 1 <= month <= 12:
                    mapping[file_hash] = f"Relevé {MONTH_NAMES_FR[month - 1]} {year}"
    except Exception:
        pass
    return mapping


def _enrich_releve_name(name: str, releve_map: dict) -> str:
    """Remplace pdf_HASH.pdf par 'Relevé Mois Année.pdf' si possible."""
    basename = name.split("/")[-1]
    m = re.match(r"pdf_([a-f0-9]+)\.pdf", basename)
    if m and m.group(1) in releve_map:
        display = releve_map[m.group(1)]
        folder = name.rsplit("/", 1)[0] if "/" in name else ""
        new_basename = f"{display}.pdf"
        return f"{folder}/{new_basename}" if folder else new_basename
    return name


def list_zip_contents(filename: str) -> list:
    """Liste les fichiers contenus dans un ZIP d'export avec noms enrichis."""
    path = EXPORTS_DIR / filename
    if not path.exists() or path.suffix != ".zip":
        return []
    try:
        releve_map = _build_releve_display_map()
        with zipfile.ZipFile(path, "r") as zf:
            entries = []
            for info in zf.infolist():
                if info.is_dir():
                    continue
                display_name = _enrich_releve_name(info.filename, releve_map)
                entries.append({
                    "name": display_name,
                    "size": info.file_size,
                    "size_human": _format_size(info.file_size),
                })
            return entries
    except Exception:
        return []


def list_exports() -> list:
    """Liste tous les exports generes."""
    ensure_directories()
    exports = []
    if not EXPORTS_DIR.exists():
        return exports

    for f in sorted(EXPORTS_DIR.iterdir(), reverse=True):
        if f.suffix == ".zip":
            stat = f.stat()
            year, month, month_name = _parse_export_filename(f.name)
            exports.append({
                "filename": f.name,
                "year": year,
                "month": month,
                "month_name": month_name or "",
                "size": stat.st_size,
                "size_human": _format_size(stat.st_size),
                "created": datetime.fromtimestamp(stat.st_mtime).isoformat(),
            })

    return exports


def _parse_export_filename(filename: str) -> tuple:
    """Parse le nom d'export (nouveau, intermédiaire ou ancien format)."""
    # Nouveau format: Export_Comptable_2025_Janvier...
    match = re.match(r"Export_Comptable_(\d{4})_([A-Za-zéèêûôàù]+)", filename)
    if match:
        mois_name = match.group(2).capitalize()
        month_num = None
        for i, m in enumerate(MOIS_FR):
            if m.capitalize() == mois_name:
                month_num = i + 1
                break
        return int(match.group(1)), month_num, mois_name
    # Format intermédiaire: Export_Comptable_2025-01_Janvier_...
    match = re.match(r"Export_Comptable_(\d{4})-(\d{2})_(\w+)", filename)
    if match:
        return int(match.group(1)), int(match.group(2)), match.group(3)
    # Ancien format: export_2024_09_Septembre_...
    match = re.match(r"export_(\d{4})_(\d{2})_(\w+)", filename)
    if match:
        return int(match.group(1)), int(match.group(2)), match.group(3)
    return None, None, None


# ─── Generation d'export ───

def generate_export(
    year: int,
    month: int,
    include_csv: bool = True,
    include_pdf: bool = False,
    include_excel: bool = False,  # kept for backward compat, ignored
    include_bank_statement: bool = True,
    include_justificatifs: bool = True,
    include_reports: bool = False,
    include_compte_attente: bool = True,
) -> dict:
    """
    Genere un export ZIP complet pour un mois donne.
    """
    ensure_directories()

    op_files = operation_service.list_operation_files()
    target_file = None
    for f in op_files:
        if f.get("year") == year and f.get("month") == month:
            target_file = f
            break

    if not target_file:
        raise ValueError(f"Aucune operation trouvee pour {month:02d}/{year}")

    operations = operation_service.load_operations(target_file["filename"])
    if not operations:
        raise ValueError(f"Fichier vide pour {month:02d}/{year}")

    month_name = MOIS_FR[month - 1].capitalize() if 1 <= month <= 12 else str(month)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    zip_name = f"Export_Comptable_{year}_{month_name}_{timestamp}.zip"
    zip_path = EXPORTS_DIR / zip_name

    # Preparer les operations classees
    prepared = _prepare_export_operations(operations, target_file["filename"])

    files_included = []

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:

        # ── CSV ──
        if include_csv:
            csv_content = _generate_csv_content(prepared, month_name, year)
            arc_name = _export_filename(year, month, "csv")
            zf.writestr(arc_name, csv_content)
            files_included.append({"name": arc_name, "type": "csv"})

        # ── PDF ──
        if include_pdf:
            pdf_bytes = _generate_pdf_content(prepared, month_name, year, month)
            arc_name = _export_filename(year, month, "pdf")
            zf.writestr(arc_name, pdf_bytes)
            files_included.append({"name": arc_name, "type": "pdf"})

        # ── Releve bancaire ──
        if include_bank_statement:
            bank_pdf = _find_bank_statement(target_file["filename"])
            if bank_pdf and bank_pdf.exists():
                arc_name = f"releve_bancaire_{month_name}_{year}.pdf"
                zf.write(str(bank_pdf), arcname=arc_name)
                files_included.append({"name": arc_name, "type": "bank_pdf"})

        # ── Justificatifs ──
        if include_justificatifs:
            just_count = 0
            for op in operations:
                lien = op.get("Lien justificatif", "")
                if lien:
                    just_filename = Path(lien).name
                    just_path = JUSTIFICATIFS_TRAITES_DIR / just_filename
                    if just_path.exists():
                        arc_name = f"justificatifs/{just_filename}"
                        zf.write(str(just_path), arcname=arc_name)
                        just_count += 1
            if just_count > 0:
                files_included.append({"name": f"justificatifs/ ({just_count} fichiers)", "type": "justificatifs"})

        # ── Rapports existants ──
        if include_reports:
            reports_added = _add_existing_reports(zf, year, month, month_name)
            for r in reports_added:
                files_included.append({"name": r, "type": "report"})

        # ── Compte d'attente ──
        if include_compte_attente:
            try:
                from backend.services import alerte_export_service
                for ca_fmt in ("pdf", "csv"):
                    ca_result = alerte_export_service.export_compte_attente(year, month, ca_fmt)
                    ca_filename = ca_result["filename"]
                    ca_path = EXPORTS_DIR / ca_filename
                    if ca_path.exists():
                        arc_name = f"compte_attente/{ca_filename}"
                        zf.write(str(ca_path), arcname=arc_name)
                        files_included.append({"name": arc_name, "type": "compte_attente"})
            except Exception as e:
                logger.warning("Impossible de generer le compte d'attente: %s", e)

    totals = prepared["totals"]

    return {
        "filename": zip_name,
        "year": year,
        "month": month,
        "month_name": month_name,
        "size": zip_path.stat().st_size,
        "size_human": _format_size(zip_path.stat().st_size),
        "operations_count": len(operations),
        "total_debit": totals["charges_pro"],
        "total_credit": totals["recettes_pro"],
        "solde": totals["solde_bnc"],
        "justificatif_count": sum(1 for op in operations if op.get("Justificatif")),
        "files_included": files_included,
        "created": datetime.now().isoformat(),
    }


# ─── Contenu CSV ───

def _generate_csv_content(prepared: dict, month_name: str, year: int) -> str:
    """
    Genere le contenu CSV avec separateur ;, format FR, sections pro/perso/attente.
    UTF-8 BOM + CRLF.
    """
    lines = []

    columns = ["Date", "Libellé", "Débit", "Crédit", "Catégorie",
               "Sous-catégorie", "Justificatif", "Commentaire"]

    lines.append(";".join(columns))

    def _csv_row(line: dict) -> str:
        debit = _format_amount_fr(line["Debit"]) if line["Debit"] > 0 else "0,00"
        credit = _format_amount_fr(line["Credit"]) if line["Credit"] > 0 else "0,00"
        libelle = line.get("Libelle", "")
        # Escape ; in fields
        if ";" in libelle:
            libelle = f'"{libelle}"'
        comment = line.get("Commentaire", "") or ""
        if ";" in comment:
            comment = f'"{comment}"'
        return ";".join([
            line.get("Date", ""),
            libelle,
            debit,
            credit,
            line.get("Categorie", ""),
            line.get("Sous_categorie", ""),
            line.get("Justificatif", ""),
            comment,
        ])

    # Section professionnelle
    for line in prepared["pro"]:
        lines.append(_csv_row(line))

    totals = prepared["totals"]

    # Ligne vide + total pro
    lines.append("")
    lines.append(";".join([
        "", "",
        _format_amount_fr(totals["charges_pro"]),
        _format_amount_fr(totals["recettes_pro"]),
        "", "", "", "TOTAL PROFESSIONNEL",
    ]))

    # Section perso
    if prepared["perso"]:
        lines.append("")
        for line in prepared["perso"]:
            lines.append(_csv_row(line))
        lines.append(";".join([
            "", "",
            _format_amount_fr(totals["debit_perso"]),
            _format_amount_fr(totals["credit_perso"]),
            "", "", "",
            f"MOUVEMENTS PERSONNELS EXCLUS ({totals['nb_perso']} opérations)",
        ]))

    # Section attente
    if prepared["attente"]:
        lines.append("")
        for line in prepared["attente"]:
            lines.append(_csv_row(line))
        lines.append(";".join([
            "", "",
            _format_amount_fr(totals["debit_attente"]),
            _format_amount_fr(totals["credit_attente"]),
            "", "", "",
            f"COMPTE D'ATTENTE ({totals['nb_attente']} opérations)",
        ]))

    # Joindre avec CRLF + BOM
    content = "\r\n".join(lines) + "\r\n"
    return "\ufeff" + content


# ─── Contenu PDF ───

def _generate_pdf_content(prepared: dict, month_name: str, year: int, month: int) -> bytes:
    """Genere un PDF professionnel avec logo, sections, totaux, pagination."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import (
        BaseDocTemplate, PageTemplate, Frame,
        Table, TableStyle, Paragraph, Spacer, Image,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER

    page_w, page_h = landscape(A4)
    margin_lr = 1.5 * cm
    margin_tb = 1.5 * cm

    # ── Styles ──
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ExportTitle", parent=styles["Heading1"],
        fontSize=16, spaceAfter=2, alignment=TA_CENTER,
    )
    sub_style = ParagraphStyle(
        "ExportSub", parent=styles["Normal"],
        fontSize=10, spaceAfter=2, textColor=colors.HexColor("#666666"),
        alignment=TA_CENTER,
    )
    date_style = ParagraphStyle(
        "ExportDate", parent=styles["Normal"],
        fontSize=9, spaceAfter=10, textColor=colors.HexColor("#999999"),
        alignment=TA_CENTER,
    )
    cell_style = ParagraphStyle(
        "Cell", parent=styles["Normal"], fontSize=7, leading=9,
    )
    cell_right = ParagraphStyle(
        "CellR", parent=styles["Normal"], fontSize=7, leading=9, alignment=TA_RIGHT,
    )
    section_style = ParagraphStyle(
        "Section", parent=styles["Normal"], fontSize=9, leading=11,
        textColor=colors.HexColor("#333333"),
    )
    recap_style = ParagraphStyle(
        "Recap", parent=styles["Normal"], fontSize=9, leading=13,
    )
    recap_bold = ParagraphStyle(
        "RecapB", parent=styles["Normal"], fontSize=10, leading=13,
        fontName="Helvetica-Bold",
    )

    # ── Column widths ──
    usable_w = page_w - 2 * margin_lr
    col_date = 70
    col_debit = 80
    col_credit = 80
    col_cat = 80
    col_subcat = 70
    col_just = 100
    col_libelle = usable_w - col_date - col_debit - col_credit - col_cat - col_subcat - col_just
    col_widths = [col_date, col_libelle, col_debit, col_credit, col_cat, col_subcat, col_just]

    # ── Footer callback ──
    footer_info = {"month_name": month_name, "year": year}

    def _footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setStrokeColor(colors.HexColor("#CCCCCC"))
        canvas.line(margin_lr, margin_tb - 8 * mm, page_w - margin_lr, margin_tb - 8 * mm)
        canvas.drawString(margin_lr, margin_tb - 12 * mm, f"Page {doc.page}")
        right_text = f"{APP_NAME} — {footer_info['month_name']} {footer_info['year']}"
        canvas.drawRightString(page_w - margin_lr, margin_tb - 12 * mm, right_text)
        canvas.restoreState()

    # ── Build doc ──
    buffer = io.BytesIO()
    doc = BaseDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=margin_lr, rightMargin=margin_lr,
        topMargin=margin_tb, bottomMargin=margin_tb + 5 * mm,
    )
    frame = Frame(
        margin_lr, margin_tb, usable_w, page_h - 2 * margin_tb - 5 * mm,
        id="main",
    )
    doc.addPageTemplates([PageTemplate(id="main", frames=[frame], onPage=_footer)])

    elements = []

    # ── Header with logo ──
    logo_path = ASSETS_DIR / "logo_lockup_light_400.png"
    header_parts = []
    if logo_path.exists():
        try:
            logo = Image(str(logo_path), width=120, height=40)
            logo.hAlign = "LEFT"
            header_parts.append(logo)
            elements.append(logo)
            elements.append(Spacer(1, 6))
        except Exception:
            pass

    elements.append(Paragraph(f"Export Comptable", title_style))
    elements.append(Paragraph(f"{month_name} {year}", sub_style))
    elements.append(Paragraph(
        f"Généré le {datetime.now().strftime('%d/%m/%Y')}", date_style
    ))
    elements.append(Spacer(1, 8))

    totals = prepared["totals"]

    # ── Section helper ──
    def _section_header_row(text: str) -> list:
        """Returns a full-width section header for the table."""
        return [Paragraph(f"<b>{text}</b>", section_style),
                "", "", "", "", "", ""]

    def _build_data_rows(ops: list) -> list:
        """Convert normalized op dicts to table rows."""
        rows = []
        for idx, line in enumerate(ops):
            debit_str = _format_amount_fr(line["Debit"]) if line["Debit"] > 0 else ""
            credit_str = _format_amount_fr(line["Credit"]) if line["Credit"] > 0 else ""
            rows.append([
                line.get("Date", ""),
                Paragraph(str(line.get("Libelle", ""))[:90], cell_style),
                Paragraph(debit_str, cell_right),
                Paragraph(credit_str, cell_right),
                str(line.get("Categorie", "")),
                str(line.get("Sous_categorie", "")),
                str(line.get("Justificatif", ""))[:20],
            ])
        return rows

    def _total_row(label: str, debit: float, credit: float) -> list:
        return [
            Paragraph(f"<b>{label}</b>", cell_style),
            "",
            Paragraph(f"<b>{_format_amount_fr(debit)}</b>", cell_right),
            Paragraph(f"<b>{_format_amount_fr(credit)}</b>", cell_right),
            "", "", "",
        ]

    # ── Build main table data ──
    headers = ["Date", "Libellé", "Débit", "Crédit", "Catégorie", "Sous-cat.", "Justificatif"]
    table_data = [headers]

    # Section PRO header
    table_data.append(_section_header_row("OPÉRATIONS PROFESSIONNELLES"))
    pro_start = len(table_data)
    table_data.extend(_build_data_rows(prepared["pro"]))
    pro_end = len(table_data)
    table_data.append(_total_row("TOTAL PROFESSIONNEL", totals["charges_pro"], totals["recettes_pro"]))
    total_pro_row = len(table_data) - 1

    # Section PERSO
    perso_header_row = None
    total_perso_row = None
    if prepared["perso"]:
        table_data.append(_section_header_row(
            f"MOUVEMENTS PERSONNELS EXCLUS ({totals['nb_perso']} opérations)"
        ))
        perso_header_row = len(table_data) - 1
        perso_start = len(table_data)
        table_data.extend(_build_data_rows(prepared["perso"]))
        table_data.append(_total_row("TOTAL PERSO", totals["debit_perso"], totals["credit_perso"]))
        total_perso_row = len(table_data) - 1

    # Section ATTENTE
    attente_header_row = None
    total_attente_row = None
    if prepared["attente"]:
        table_data.append(_section_header_row(
            f"COMPTE D'ATTENTE ({totals['nb_attente']} opérations)"
        ))
        attente_header_row = len(table_data) - 1
        attente_start = len(table_data)
        table_data.extend(_build_data_rows(prepared["attente"]))
        table_data.append(_total_row("TOTAL ATTENTE", totals["debit_attente"], totals["credit_attente"]))
        total_attente_row = len(table_data) - 1

    # ── Table style ──
    style_cmds = [
        # Header row
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#811971")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        # Global
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        # Amounts right-aligned
        ("ALIGN", (2, 1), (3, -1), "RIGHT"),
        # Section header: PRO
        ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#D5E8F0")),
        ("SPAN", (0, 1), (-1, 1)),
        # Total PRO
        ("BACKGROUND", (0, total_pro_row), (-1, total_pro_row), colors.HexColor("#E8E8E8")),
    ]

    # Alternating row colors for PRO data rows
    for i in range(pro_start, pro_end):
        if (i - pro_start) % 2 == 0:
            style_cmds.append(("BACKGROUND", (0, i), (-1, i), colors.white))
        else:
            style_cmds.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#F8F8F8")))

    # Section headers and totals for perso
    if perso_header_row is not None:
        style_cmds.append(("BACKGROUND", (0, perso_header_row), (-1, perso_header_row), colors.HexColor("#D5E8F0")))
        style_cmds.append(("SPAN", (0, perso_header_row), (-1, perso_header_row)))
    if total_perso_row is not None:
        style_cmds.append(("BACKGROUND", (0, total_perso_row), (-1, total_perso_row), colors.HexColor("#E8E8E8")))
        # Alternating for perso data
        perso_data_start = perso_header_row + 1 if perso_header_row is not None else 0
        for i in range(perso_data_start, total_perso_row):
            if (i - perso_data_start) % 2 == 0:
                style_cmds.append(("BACKGROUND", (0, i), (-1, i), colors.white))
            else:
                style_cmds.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#F8F8F8")))

    # Section headers and totals for attente
    if attente_header_row is not None:
        style_cmds.append(("BACKGROUND", (0, attente_header_row), (-1, attente_header_row), colors.HexColor("#D5E8F0")))
        style_cmds.append(("SPAN", (0, attente_header_row), (-1, attente_header_row)))
    if total_attente_row is not None:
        style_cmds.append(("BACKGROUND", (0, total_attente_row), (-1, total_attente_row), colors.HexColor("#E8E8E8")))
        attente_data_start = attente_header_row + 1 if attente_header_row is not None else 0
        for i in range(attente_data_start, total_attente_row):
            if (i - attente_data_start) % 2 == 0:
                style_cmds.append(("BACKGROUND", (0, i), (-1, i), colors.white))
            else:
                style_cmds.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#F8F8F8")))

    ops_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    ops_table.setStyle(TableStyle(style_cmds))
    elements.append(ops_table)

    # ── Recapitulatif ──
    elements.append(Spacer(1, 16))

    recap_data = [
        [Paragraph("<b>RÉCAPITULATIF</b>", recap_bold), ""],
        [Paragraph("Recettes professionnelles :", recap_style),
         Paragraph(f"<b>{_format_amount_fr(totals['recettes_pro'])} €</b>", ParagraphStyle("RR", parent=recap_style, alignment=TA_RIGHT))],
        [Paragraph("Charges professionnelles :", recap_style),
         Paragraph(f"<b>{_format_amount_fr(totals['charges_pro'])} €</b>", ParagraphStyle("RR2", parent=recap_style, alignment=TA_RIGHT))],
        [Paragraph("<b>Solde BNC :</b>", recap_bold),
         Paragraph(f"<b>{_format_amount_fr(totals['solde_bnc'])} €</b>", ParagraphStyle("RR3", parent=recap_bold, alignment=TA_RIGHT))],
    ]
    if totals["nb_perso"] > 0:
        recap_data.append([
            Paragraph(f"Mouvements personnels exclus ({totals['nb_perso']} ops) :", recap_style),
            Paragraph(f"{_format_amount_fr(totals['total_perso'])} €", ParagraphStyle("RR4", parent=recap_style, alignment=TA_RIGHT)),
        ])
    if totals["nb_attente"] > 0:
        recap_data.append([
            Paragraph(f"En attente de ventilation ({totals['nb_attente']} ops) :", recap_style),
            Paragraph(f"{_format_amount_fr(totals['total_attente'])} €", ParagraphStyle("RR5", parent=recap_style, alignment=TA_RIGHT)),
        ])

    recap_table = Table(recap_data, colWidths=[usable_w * 0.55, usable_w * 0.45])
    recap_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F0F4F8")),
        ("SPAN", (0, 0), (-1, 0)),
        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#FAFAFA")),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("LINEBELOW", (0, 0), (-1, 0), 0.5, colors.HexColor("#CCCCCC")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
    ]))
    elements.append(recap_table)

    doc.build(elements)
    return buffer.getvalue()


# ─── Ventilation par catégorie (PDF + CSV avec sous-totaux) ───

def _group_by_category(prepared: dict) -> dict:
    """Agrège toutes les lignes (pro + perso + attente) par (Catégorie, Sous-catégorie).

    Ventilations déjà explosées par `_prepare_export_operations`. Catégorie vide
    ou "Ventilé" parent jamais présent ici (les enfants arrivent avec leur propre
    cat/sous-cat). Retour trié par cat alpha, sub alpha.
    """
    all_lines = list(prepared.get("pro") or []) + list(prepared.get("perso") or []) + list(prepared.get("attente") or [])

    groups: dict[str, dict] = {}
    for line in all_lines:
        cat = (line.get("Categorie") or "").strip() or "(Sans catégorie)"
        sub = (line.get("Sous_categorie") or "").strip() or "(Sans sous-catégorie)"
        g = groups.setdefault(cat, {"lines_by_sub": {}, "debit": 0.0, "credit": 0.0, "count": 0})
        sg = g["lines_by_sub"].setdefault(sub, {"lines": [], "debit": 0.0, "credit": 0.0})
        sg["lines"].append(line)
        sg["debit"] += _safe_float(line.get("Debit"))
        sg["credit"] += _safe_float(line.get("Credit"))
        g["debit"] += _safe_float(line.get("Debit"))
        g["credit"] += _safe_float(line.get("Credit"))
        g["count"] += 1

    ordered = []
    for cat in sorted(groups.keys(), key=lambda s: s.lower()):
        g = groups[cat]
        subcats = []
        for sub in sorted(g["lines_by_sub"].keys(), key=lambda s: s.lower()):
            sg = g["lines_by_sub"][sub]
            sg["lines"].sort(key=lambda l: l.get("Date") or "")
            subcats.append({
                "sous_categorie": sub,
                "lines": sg["lines"],
                "debit": sg["debit"],
                "credit": sg["credit"],
            })
        ordered.append({
            "categorie": cat,
            "subcats": subcats,
            "debit": g["debit"],
            "credit": g["credit"],
            "count": g["count"],
            "net": g["debit"] - g["credit"],
        })

    grand_debit = sum(g["debit"] for g in ordered)
    grand_credit = sum(g["credit"] for g in ordered)
    return {
        "groups": ordered,
        "grand_debit": grand_debit,
        "grand_credit": grand_credit,
        "grand_net": grand_debit - grand_credit,
        "nb_lines": sum(g["count"] for g in ordered),
    }


def _category_csv_filename(year: int, month: int) -> str:
    month_name = MONTH_NAMES_FR[month - 1]
    return f"Ventilation_par_categorie_{year}-{month:02d}_{month_name}.csv"


def _category_pdf_filename(year: int, month: int) -> str:
    month_name = MONTH_NAMES_FR[month - 1]
    return f"Ventilation_par_categorie_{year}-{month:02d}_{month_name}.pdf"


def _generate_csv_by_category(prepared: dict, month_name: str, year: int) -> str:
    """CSV groupé par catégorie avec sous-totaux + total général. Format FR."""
    grouped = _group_by_category(prepared)
    lines = []
    lines.append(f"Ventilation par catégorie — {month_name} {year}")
    lines.append(f"{grouped['nb_lines']} opérations")
    lines.append("")
    columns = ["Date", "Libellé", "Débit", "Crédit", "Catégorie", "Sous-catégorie", "Justificatif", "Commentaire"]
    lines.append(";".join(columns))

    def _csv_row(line: dict) -> str:
        debit = _format_amount_fr(line["Debit"]) if line["Debit"] > 0 else "0,00"
        credit = _format_amount_fr(line["Credit"]) if line["Credit"] > 0 else "0,00"
        libelle = (line.get("Libelle") or "").replace('"', "'")
        if ";" in libelle:
            libelle = f'"{libelle}"'
        comment = (line.get("Commentaire") or "").replace('"', "'")
        if ";" in comment:
            comment = f'"{comment}"'
        return ";".join([
            line.get("Date", ""),
            libelle,
            debit,
            credit,
            line.get("Categorie", ""),
            line.get("Sous_categorie", ""),
            line.get("Justificatif", ""),
            comment,
        ])

    for g in grouped["groups"]:
        for sg in g["subcats"]:
            for ln in sg["lines"]:
                lines.append(_csv_row(ln))
            # Sous-total sous-catégorie
            lines.append(";".join([
                "", "",
                _format_amount_fr(sg["debit"]),
                _format_amount_fr(sg["credit"]),
                g["categorie"], sg["sous_categorie"], "",
                f"Sous-total {sg['sous_categorie']} ({len(sg['lines'])} op.)",
            ]))
        # Sous-total catégorie
        lines.append(";".join([
            "", "",
            _format_amount_fr(g["debit"]),
            _format_amount_fr(g["credit"]),
            g["categorie"], "", "",
            f"TOTAL {g['categorie']} ({g['count']} op.)",
        ]))
        lines.append("")

    # Total général
    lines.append(";".join([
        "", "",
        _format_amount_fr(grouped["grand_debit"]),
        _format_amount_fr(grouped["grand_credit"]),
        "", "", "",
        f"TOTAL GÉNÉRAL ({grouped['nb_lines']} op.)",
    ]))
    content = "\r\n".join(lines) + "\r\n"
    return "\ufeff" + content


def _generate_pdf_by_category(prepared: dict, month_name: str, year: int, month: int) -> bytes:
    """PDF A4 portrait : sections par catégorie, sous-sections par sous-cat, sous-totaux.

    Layout compact : titre centré, logo, tableau par catégorie avec lignes
    ops + lignes sous-totales en emphase, total général en fin de document.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm, mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    from reportlab.platypus import (
        BaseDocTemplate, PageTemplate, Frame, Table, TableStyle,
        Paragraph, Spacer, Image, KeepTogether,
    )
    import io

    grouped = _group_by_category(prepared)

    page_w, page_h = A4
    margin_lr = 1.4 * cm
    margin_tb = 1.4 * cm

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CatTitle", parent=styles["Heading1"],
        fontSize=15, spaceAfter=2, alignment=TA_CENTER,
    )
    sub_style = ParagraphStyle(
        "CatSub", parent=styles["Normal"],
        fontSize=9, spaceAfter=10, alignment=TA_CENTER,
        textColor=colors.HexColor("#666666"),
    )
    cat_header_style = ParagraphStyle(
        "CatHeader", parent=styles["Normal"],
        fontSize=11, leading=13, spaceBefore=6, spaceAfter=3,
        fontName="Helvetica-Bold",
        textColor=colors.HexColor("#1F3A68"),
    )
    subcat_style = ParagraphStyle(
        "SubCat", parent=styles["Normal"],
        fontSize=9, leading=11, spaceBefore=2, spaceAfter=1,
        fontName="Helvetica-Oblique",
        textColor=colors.HexColor("#444444"),
    )
    cell_style = ParagraphStyle("CatCell", parent=styles["Normal"], fontSize=7, leading=9)
    cell_right = ParagraphStyle("CatCellR", parent=styles["Normal"], fontSize=7, leading=9, alignment=TA_RIGHT)

    usable_w = page_w - 2 * margin_lr
    col_date = 58
    col_debit = 62
    col_credit = 62
    col_just = 90
    col_libelle = usable_w - col_date - col_debit - col_credit - col_just
    col_widths = [col_date, col_libelle, col_debit, col_credit, col_just]

    def _footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setStrokeColor(colors.HexColor("#CCCCCC"))
        canvas.line(margin_lr, margin_tb - 8 * mm, page_w - margin_lr, margin_tb - 8 * mm)
        canvas.drawString(margin_lr, margin_tb - 12 * mm, f"Page {doc.page}")
        right_text = f"Ventilation par catégorie — {month_name} {year}"
        canvas.drawRightString(page_w - margin_lr, margin_tb - 12 * mm, right_text)
        canvas.restoreState()

    buffer = io.BytesIO()
    doc = BaseDocTemplate(
        buffer, pagesize=A4,
        leftMargin=margin_lr, rightMargin=margin_lr,
        topMargin=margin_tb, bottomMargin=margin_tb + 5 * mm,
    )
    frame = Frame(
        margin_lr, margin_tb, usable_w, page_h - 2 * margin_tb - 5 * mm,
        id="main", leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0,
    )
    doc.addPageTemplates([PageTemplate(id="default", frames=[frame], onPage=_footer)])

    elements = []
    # Logo (optionnel)
    try:
        from backend.core.config import BASE_DIR
        logo_path = BASE_DIR / "backend" / "assets" / "logo_lockup_light_400.png"
        if logo_path.exists():
            img = Image(str(logo_path))
            img.drawHeight = 1.3 * cm
            img.drawWidth = 4.2 * cm
            img.hAlign = "CENTER"
            elements.append(img)
            elements.append(Spacer(1, 4))
    except Exception:
        pass

    elements.append(Paragraph(f"Ventilation par catégorie — {month_name} {year}", title_style))
    elements.append(Paragraph(f"{grouped['nb_lines']} opérations — {len(grouped['groups'])} catégories", sub_style))

    header_row = [
        Paragraph("<b>Date</b>", cell_style),
        Paragraph("<b>Libellé</b>", cell_style),
        Paragraph("<b>Débit</b>", cell_right),
        Paragraph("<b>Crédit</b>", cell_right),
        Paragraph("<b>Justificatif</b>", cell_style),
    ]

    for g in grouped["groups"]:
        elements.append(Paragraph(
            f"{g['categorie']} <font size='8' color='#888888'>· {g['count']} op.</font>",
            cat_header_style,
        ))
        for sg in g["subcats"]:
            if sg["sous_categorie"] not in ("(Sans sous-catégorie)",):
                elements.append(Paragraph(f"— {sg['sous_categorie']} ({len(sg['lines'])} op.)", subcat_style))
            data = [header_row]
            for ln in sg["lines"]:
                debit = _format_amount_fr(ln["Debit"]) if ln["Debit"] > 0 else ""
                credit = _format_amount_fr(ln["Credit"]) if ln["Credit"] > 0 else ""
                data.append([
                    Paragraph(ln.get("Date", ""), cell_style),
                    Paragraph(ln.get("Libelle", ""), cell_style),
                    Paragraph(debit, cell_right),
                    Paragraph(credit, cell_right),
                    Paragraph(ln.get("Justificatif", ""), cell_style),
                ])
            # Sous-total sous-catégorie
            data.append([
                "", Paragraph("<b>Sous-total</b>", cell_right),
                Paragraph(f"<b>{_format_amount_fr(sg['debit'])}</b>", cell_right),
                Paragraph(f"<b>{_format_amount_fr(sg['credit'])}</b>", cell_right),
                "",
            ])
            tbl = Table(data, colWidths=col_widths, repeatRows=1)
            tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D5E8F0")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1F3A68")),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#DDDDDD")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F3F6F9")),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ]))
            elements.append(tbl)
            elements.append(Spacer(1, 3))
        # Total catégorie
        cat_tbl = Table(
            [[
                Paragraph(f"<b>TOTAL {g['categorie']}</b>", cell_style),
                Paragraph(f"<b>{_format_amount_fr(g['debit'])}</b>", cell_right),
                Paragraph(f"<b>{_format_amount_fr(g['credit'])}</b>", cell_right),
                Paragraph(f"<b>Net {_format_amount_fr(g['net'])}</b>", cell_right),
            ]],
            colWidths=[col_libelle + col_date, col_debit, col_credit, col_just],
        )
        cat_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#1F3A68")),
            ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(cat_tbl)
        elements.append(Spacer(1, 10))

    # Total général
    grand = Table(
        [[
            Paragraph("<b>TOTAL GÉNÉRAL</b>", cell_style),
            Paragraph(f"<b>{_format_amount_fr(grouped['grand_debit'])}</b>", cell_right),
            Paragraph(f"<b>{_format_amount_fr(grouped['grand_credit'])}</b>", cell_right),
            Paragraph(f"<b>Net {_format_amount_fr(grouped['grand_net'])}</b>", cell_right),
        ]],
        colWidths=[col_libelle + col_date, col_debit, col_credit, col_just],
    )
    grand.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#111827")),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(grand)

    doc.build(elements)
    return buffer.getvalue()


# ─── Contenu Excel ───

def _generate_excel_content(prepared: dict, month_name: str, year: int) -> bytes:
    """Genere un fichier Excel avec les 3 sections."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Opérations"

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="811971", end_color="811971", fill_type="solid")
    section_fill = PatternFill(start_color="D5E8F0", end_color="D5E8F0", fill_type="solid")
    total_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    total_font = Font(bold=True, size=10)
    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    columns = ["Date", "Libellé", "Débit", "Crédit", "Catégorie",
               "Sous-catégorie", "Justificatif", "Commentaire"]

    # Header row
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    row_idx = 2

    def _write_section_header(ws, row, text):
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(bold=True, size=10)
        for c in range(1, 9):
            ws.cell(row=row, column=c).fill = section_fill
            ws.cell(row=row, column=c).border = thin_border
        return row + 1

    def _write_ops(ws, row, ops):
        for line in ops:
            ws.cell(row=row, column=1, value=line.get("Date", "")).border = thin_border
            ws.cell(row=row, column=2, value=line.get("Libelle", "")).border = thin_border
            c3 = ws.cell(row=row, column=3, value=line["Debit"] if line["Debit"] > 0 else 0)
            c3.number_format = '#,##0.00'
            c3.alignment = Alignment(horizontal="right")
            c3.border = thin_border
            c4 = ws.cell(row=row, column=4, value=line["Credit"] if line["Credit"] > 0 else 0)
            c4.number_format = '#,##0.00'
            c4.alignment = Alignment(horizontal="right")
            c4.border = thin_border
            ws.cell(row=row, column=5, value=line.get("Categorie", "")).border = thin_border
            ws.cell(row=row, column=6, value=line.get("Sous_categorie", "")).border = thin_border
            ws.cell(row=row, column=7, value=line.get("Justificatif", "")).border = thin_border
            ws.cell(row=row, column=8, value=line.get("Commentaire", "")).border = thin_border
            row += 1
        return row

    def _write_total_row(ws, row, label, debit, credit):
        ws.cell(row=row, column=1, value=label).font = total_font
        for c in range(1, 9):
            ws.cell(row=row, column=c).fill = total_fill
            ws.cell(row=row, column=c).border = thin_border
        c3 = ws.cell(row=row, column=3, value=debit)
        c3.number_format = '#,##0.00'
        c3.alignment = Alignment(horizontal="right")
        c3.font = total_font
        c4 = ws.cell(row=row, column=4, value=credit)
        c4.number_format = '#,##0.00'
        c4.alignment = Alignment(horizontal="right")
        c4.font = total_font
        return row + 1

    totals = prepared["totals"]

    # PRO section
    row_idx = _write_section_header(ws, row_idx, "OPÉRATIONS PROFESSIONNELLES")
    row_idx = _write_ops(ws, row_idx, prepared["pro"])
    row_idx = _write_total_row(ws, row_idx, "TOTAL PROFESSIONNEL", totals["charges_pro"], totals["recettes_pro"])
    row_idx += 1  # blank row

    # PERSO section
    if prepared["perso"]:
        row_idx = _write_section_header(ws, row_idx, f"MOUVEMENTS PERSONNELS EXCLUS ({totals['nb_perso']} opérations)")
        row_idx = _write_ops(ws, row_idx, prepared["perso"])
        row_idx = _write_total_row(ws, row_idx, "TOTAL PERSO", totals["debit_perso"], totals["credit_perso"])
        row_idx += 1

    # ATTENTE section
    if prepared["attente"]:
        row_idx = _write_section_header(ws, row_idx, f"COMPTE D'ATTENTE ({totals['nb_attente']} opérations)")
        row_idx = _write_ops(ws, row_idx, prepared["attente"])
        row_idx = _write_total_row(ws, row_idx, "TOTAL ATTENTE", totals["debit_attente"], totals["credit_attente"])

    col_widths = [12, 45, 12, 12, 18, 18, 18, 25]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ─── Releve bancaire ───

def _find_bank_statement(operation_filename: str) -> Optional[Path]:
    """Trouve le PDF du releve bancaire associe au fichier d'operations.

    Cas géré :
    1. Fichier original `operations_YYYYMMDD_HHMMSS_<hex>.json` → `pdf_<hex>.pdf` direct.
    2. Fichier issu d'un merge/split `operations_(merged|split)_YYYYMM_*.json` → le
       hex d'origine est perdu ; on scanne `_archive/` pour retrouver un source
       archivé (`operations_YYYYMMDD_HHMMSS_<hex>.json.bak_*`) dont le mois dominant
       correspond et dont la période recouvre `YYYYMM`, puis on résout son `pdf_<hex>.pdf`.
    """
    # Cas 1 : format original
    match = re.search(r"operations_\d{8}_\d{6}_([a-f0-9]+)\.json", operation_filename)
    if match:
        file_id = match.group(1)
        pdf_path = IMPORTS_RELEVES_DIR / f"pdf_{file_id}.pdf"
        if pdf_path.exists():
            return pdf_path

    # Cas 2 : merged/split → scan _archive/ pour trouver le source d'origine.
    #
    # Stratégie : parcourir les archives `operations_YYYYMMDD_HHMMSS_<hex>.json.bak_*`,
    # compter combien d'ops tombent dans le mois cible, et retourner le relevé de
    # l'archive avec le plus d'ops dans ce mois (≥ 3 pour éviter les faux positifs
    # des merged multi-mois). Résout les 2 cas :
    # - merged : plusieurs archives (un par mois pré-existant), chacune a 100% du mois
    # - split  : une seule archive (fichier multi-mois) avec sous-ensemble du mois cible
    merge_split = re.search(r"operations_(?:merged|split)_(\d{4})(\d{2})_", operation_filename)
    if not merge_split:
        return None
    target_year = int(merge_split.group(1))
    target_month = int(merge_split.group(2))

    from backend.core.config import IMPORTS_OPERATIONS_DIR
    archive_dir = IMPORTS_OPERATIONS_DIR / "_archive"
    if not archive_dir.exists():
        return None

    # Tri priorisant les fichiers mensuels (concentration élevée) sur les multi-mois.
    # Score = (is_monthly, concentration_ratio, nb_in_month) — plus haut est meilleur.
    # is_monthly = 1 si ≥80% des ops du fichier sont dans le mois cible, sinon 0.
    best_match: Optional[tuple[tuple[int, float, int], Path]] = None
    for archived in archive_dir.glob("operations_*_*.json.bak_*"):
        m = re.search(r"operations_\d{8}_\d{6}_([a-f0-9]+)\.json\.bak_", archived.name)
        if not m:
            continue
        hex_id = m.group(1)
        pdf_candidate = IMPORTS_RELEVES_DIR / f"pdf_{hex_id}.pdf"
        if not pdf_candidate.exists():
            continue
        try:
            with open(archived, "r", encoding="utf-8") as f:
                ops = json.load(f)
        except (OSError, json.JSONDecodeError):
            continue
        if not isinstance(ops, list) or not ops:
            continue
        nb_in_month = 0
        for op in ops:
            d = (op.get("Date") or "").strip()
            if len(d) < 7:
                continue
            try:
                if int(d[:4]) == target_year and int(d[5:7]) == target_month:
                    nb_in_month += 1
            except ValueError:
                continue
        if nb_in_month < 3:
            continue
        total = len(ops)
        concentration = nb_in_month / total if total else 0.0
        is_monthly = 1 if concentration >= 0.8 else 0
        # Priorité : fichier mensuel ciblé > plus d'ops dans le mois > concentration
        # (évite les files qui ont juste 3-5 ops en débordement d'un autre mois)
        score = (is_monthly, nb_in_month, concentration)
        if best_match is None or score > best_match[0]:
            best_match = (score, pdf_candidate)

    return best_match[1] if best_match else None


# ─── Rapports existants ───

def _add_existing_reports(zf: zipfile.ZipFile, year: int, month: int, month_name: str) -> list:
    """Ajoute les rapports existants pour ce mois au ZIP."""
    added = []
    for f in _find_existing_reports(year, month, month_name):
        arc_name = f"rapports/{f.name}"
        zf.write(str(f), arcname=arc_name)
        added.append(arc_name)
    return added


# ─── Suppression ───

def delete_export(filename: str) -> bool:
    """Supprime un export."""
    path = EXPORTS_DIR / filename
    if path.exists() and path.suffix == ".zip":
        path.unlink()
        return True
    return False


# ─── Download path ───

def delete_single_file(filename: str) -> bool:
    """Supprime un fichier d'export individuel (PDF/CSV)."""
    path = EXPORTS_DIR / filename
    if path.exists() and path.suffix in (".pdf", ".csv", ".zip"):
        path.unlink()
        return True
    return False


def get_export_path(filename: str) -> Optional[Path]:
    """Retourne le chemin absolu d'un export (ZIP, PDF ou CSV)."""
    path = EXPORTS_DIR / filename
    if path.exists():
        return path
    return None


# ─── Utils ───

def _format_size(size_bytes: int) -> str:
    """Formate la taille en bytes lisible."""
    if size_bytes < 1024:
        return f"{size_bytes} o"
    elif size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.0f} Ko"
    else:
        return f"{size_bytes / (1024 * 1024):.1f} Mo"
