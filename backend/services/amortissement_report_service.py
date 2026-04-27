"""
Rendu des rapports amortissements multi-format (Prompt B3).

Deux fonctions de rendu unifiées consommées par les templates Rapports V2 :
  - `render_registre(year, output_path, format, filters)` : registre complet des immos
  - `render_dotations(year, output_path, format, filters)` : tableau dotations exercice

Chaque fonction dispatche selon le format (`pdf` / `csv` / `xlsx`).
Colonne `Origine` (NeuronX / Reprise {year}) intégrée — badge ambre dans le PDF
pour les immos reprises (`exercice_entree_neuronx is not None`).

`generate_dotation_pdf(year, output_path)` est conservé comme wrapper backward-compat
B1 pointant vers `render_dotations(year, path, "pdf", {"poste": "all"})`.
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Optional

from backend.core.config import ASSETS_DIR

logger = logging.getLogger(__name__)


# ─── Helpers communs ────────────────────────────────────────────────


def _fr_euro(montant: Optional[float]) -> str:
    """1 234,56 €"""
    if montant is None:
        return "0,00 €"
    formatted = f"{montant:,.2f}"
    return formatted.replace(",", " ").replace(".", ",") + " €"


def _fr_decimal(montant: Optional[float]) -> str:
    """1234,56 — sans symbole € (CSV/XLSX)."""
    if montant is None:
        return "0,00"
    return f"{montant:.2f}".replace(".", ",")


def _format_date_fr(iso: str) -> str:
    """2024-03-15 → 15/03/2024"""
    if not iso or len(iso) < 10:
        return iso or ""
    return f"{iso[8:10]}/{iso[5:7]}/{iso[0:4]}"


def _describe_filters(filters: dict) -> str:
    """Concatène les filtres actifs pour le sous-titre PDF."""
    parts = []
    statut = filters.get("statut", "all")
    if statut and statut != "all":
        parts.append(f"statut : {statut}")
    poste = filters.get("poste", "all")
    if poste and poste != "all":
        parts.append(f"poste : {poste}")
    return ", ".join(parts)


def _origine_label(immo) -> str:
    """Retourne 'NeuronX' ou 'Reprise {year}' selon `exercice_entree_neuronx`."""
    if isinstance(immo, dict):
        year_entree = immo.get("exercice_entree_neuronx")
    else:
        year_entree = getattr(immo, "exercice_entree_neuronx", None)
    if year_entree:
        return f"Reprise {year_entree}"
    return "NeuronX"


def _is_reprise(immo) -> bool:
    if isinstance(immo, dict):
        return immo.get("exercice_entree_neuronx") is not None
    return getattr(immo, "exercice_entree_neuronx", None) is not None


# ═══════════════════════════════════════════════════════════════════
# REGISTRE
# ═══════════════════════════════════════════════════════════════════


def render_registre(year: int, output_path: Path, format: str, filters: dict) -> Path:
    """Registre complet des immobilisations (actives + amorties + sorties).

    Filtres : `year` (req, contexte titre), `statut` (all|en_cours|amorti|sorti),
    `poste` (all|<poste>).
    """
    from backend.services import amortissement_service

    immos = amortissement_service.list_immobilisations_enriched(year=year)
    immos = _apply_registre_filters(immos, filters)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if format == "pdf":
        return _render_registre_pdf(immos, year, output_path, filters)
    if format == "csv":
        return _render_registre_csv(immos, year, output_path, filters)
    if format in ("xlsx", "excel"):
        return _render_registre_xlsx(immos, year, output_path, filters)
    raise ValueError(f"Format non supporté pour registre : {format}")


def _apply_registre_filters(immos: list, filters: dict) -> list:
    statut = (filters.get("statut") or "all").lower()
    poste = (filters.get("poste") or "all").lower()
    if statut != "all":
        immos = [i for i in immos if (i.get("statut") or "").lower() == statut]
    if poste != "all":
        immos = [i for i in immos if (i.get("poste") or "").lower() == poste]
    return immos


def _render_registre_pdf(immos: list, year: int, output_path: Path, filters: dict) -> Path:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Image,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=landscape(A4),  # paysage pour accueillir Origine + Cumul
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )

    styles = getSampleStyleSheet()
    story: list = []

    # Logo
    logo_path = ASSETS_DIR / "logo_lockup_light_400.png"
    if logo_path.exists():
        story.append(Image(str(logo_path), width=55 * mm, height=14 * mm))
        story.append(Spacer(1, 4 * mm))

    # Titre + sous-filtres
    filtre_label = _describe_filters(filters)
    title_text = f"Registre des immobilisations — Exercice {year}"
    if filtre_label:
        title_text += f" · {filtre_label}"
    story.append(Paragraph(title_text, styles["Title"]))
    story.append(Spacer(1, 3 * mm))

    headers = [
        "Désignation",
        "Origine",
        "Acquis le",
        "Statut",
        "Durée",
        "Base",
        "Cumul amort.",
        "VNC actuelle",
        "Poste",
    ]
    data: list[list[str]] = [headers]

    if not immos:
        data.append(["Aucune immobilisation", "", "", "", "", "", "", "", ""])
    else:
        for immo in immos:
            designation = immo.get("designation", "") or ""
            if len(designation) > 35:
                designation = designation[:35] + "…"
            base = float(immo.get("base_amortissable", 0) or 0)
            vnc_actuelle = float(immo.get("vnc_actuelle", base) or base)
            cumul = base - vnc_actuelle
            data.append([
                designation,
                _origine_label(immo),
                _format_date_fr(immo.get("date_acquisition", "")),
                immo.get("statut") or "",
                f"{immo.get('duree', 0)} ans",
                _fr_euro(base),
                _fr_euro(cumul),
                _fr_euro(vnc_actuelle),
                immo.get("poste") or "—",
            ])

        # Ligne TOTAL
        total_base = sum(float(i.get("base_amortissable", 0) or 0) for i in immos)
        total_vnc = sum(float(i.get("vnc_actuelle", 0) or 0) for i in immos)
        total_cumul = total_base - total_vnc
        data.append([
            "TOTAL",
            "",
            "",
            f"{len(immos)} immo(s)",
            "",
            _fr_euro(total_base),
            _fr_euro(total_cumul),
            _fr_euro(total_vnc),
            "",
        ])

    table = Table(
        data,
        repeatRows=1,
        colWidths=[
            55 * mm, 22 * mm, 20 * mm, 20 * mm, 14 * mm,
            26 * mm, 26 * mm, 26 * mm, 25 * mm,
        ],
    )

    style_cmds: list = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3C3489")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
        ("ALIGN", (5, 1), (7, -1), "RIGHT"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]

    if immos:
        # TOTAL row
        style_cmds += [
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#EEEDFE")),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ]
        # Coloration ligne par ligne pour Origine "Reprise" — badge ambre
        for row_idx, immo in enumerate(immos, start=1):
            if _is_reprise(immo):
                style_cmds.append(
                    ("BACKGROUND", (1, row_idx), (1, row_idx), colors.HexColor("#FAEEDA"))
                )
                style_cmds.append(
                    ("TEXTCOLOR", (1, row_idx), (1, row_idx), colors.HexColor("#854F0B"))
                )
                style_cmds.append(
                    ("FONTNAME", (1, row_idx), (1, row_idx), "Helvetica-Bold")
                )

    table.setStyle(TableStyle(style_cmds))
    story.append(table)
    story.append(Spacer(1, 6 * mm))

    # Référence légale
    ref_style = ParagraphStyle(
        "ref",
        parent=styles["Italic"],
        fontSize=8,
        textColor=colors.HexColor("#666666"),
    )
    story.append(Paragraph(
        "Référence : art. 39-1-2° du CGI, PCG art. 214-13. "
        "Régime BNC — comptabilité de recettes — amortissement linéaire.",
        ref_style,
    ))

    doc.build(story)
    logger.info(
        "PDF registre amortissements généré: %s (%d immos)",
        output_path.name, len(immos),
    )
    return output_path


def _render_registre_csv(immos: list, year: int, output_path: Path, filters: dict) -> Path:
    """CSV BOM UTF-8, séparateur `;`, virgule décimale, CRLF (Excel FR)."""
    headers = [
        "Désignation",
        "Origine",
        "Acquis le",
        "Statut",
        "Mode",
        "Durée",
        "Base amortissable",
        "Cumul amortissements",
        "VNC actuelle",
        "Quote-part (%)",
        "Poste",
        "Exercice entrée NeuronX",
    ]
    lines: list[str] = [";".join(headers)]

    for immo in immos:
        base = float(immo.get("base_amortissable", 0) or 0)
        vnc = float(immo.get("vnc_actuelle", base) or base)
        cumul = base - vnc
        year_entree = immo.get("exercice_entree_neuronx")
        lines.append(";".join([
            (immo.get("designation", "") or "").replace(";", ","),
            _origine_label(immo),
            _format_date_fr(immo.get("date_acquisition", "")),
            immo.get("statut") or "",
            immo.get("mode") or "",
            str(immo.get("duree", 0)),
            _fr_decimal(base),
            _fr_decimal(cumul),
            _fr_decimal(vnc),
            f"{float(immo.get('quote_part_pro', 100) or 100):.0f}",
            (immo.get("poste") or "").replace(";", ","),
            str(year_entree) if year_entree else "",
        ]))

    # Ligne TOTAL
    total_base = sum(float(i.get("base_amortissable", 0) or 0) for i in immos)
    total_vnc = sum(float(i.get("vnc_actuelle", 0) or 0) for i in immos)
    total_cumul = total_base - total_vnc
    lines.append(";".join([
        "TOTAL", "", "", "", "", "",
        _fr_decimal(total_base),
        _fr_decimal(total_cumul),
        _fr_decimal(total_vnc),
        "", "", "",
    ]))

    content = "\r\n".join(lines)
    output_path.write_bytes(b"\xef\xbb\xbf" + content.encode("utf-8"))
    logger.info("CSV registre amortissements généré: %s (%d immos)", output_path.name, len(immos))
    return output_path


def _render_registre_xlsx(immos: list, year: int, output_path: Path, filters: dict) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = f"Registre {year}"[:31]  # Excel sheet name limit

    headers = [
        "Désignation",
        "Origine",
        "Acquis le",
        "Statut",
        "Mode",
        "Durée",
        "Base amortissable",
        "Cumul amortissements",
        "VNC actuelle",
        "Quote-part (%)",
        "Poste",
        "Exercice entrée NeuronX",
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="3C3489", end_color="3C3489", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    reprise_fill = PatternFill(start_color="FAEEDA", end_color="FAEEDA", fill_type="solid")
    reprise_font = Font(color="854F0B", bold=True)

    for immo in immos:
        base = float(immo.get("base_amortissable", 0) or 0)
        vnc = float(immo.get("vnc_actuelle", base) or base)
        cumul = base - vnc
        year_entree = immo.get("exercice_entree_neuronx")
        row = [
            immo.get("designation", "") or "",
            _origine_label(immo),
            immo.get("date_acquisition", ""),
            immo.get("statut") or "",
            immo.get("mode") or "",
            int(immo.get("duree", 0) or 0),
            base,
            cumul,
            vnc,
            float(immo.get("quote_part_pro", 100) or 100),
            immo.get("poste") or "",
            year_entree if year_entree else "",
        ]
        ws.append(row)
        # Style Reprise sur la cellule Origine
        if _is_reprise(immo):
            cell = ws.cell(row=ws.max_row, column=2)
            cell.fill = reprise_fill
            cell.font = reprise_font

    # Format EUR sur colonnes monétaires (G/H/I = 7/8/9)
    for col_letter in ("G", "H", "I"):
        for cell in ws[col_letter][1:]:
            cell.number_format = "#,##0.00 €"
            cell.alignment = Alignment(horizontal="right")

    # Ligne TOTAL avec formules SUM
    last_row = ws.max_row
    if last_row >= 2:
        total_row = last_row + 1
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=4, value=f"{len(immos)} immo(s)")
        for col_idx, col_letter in [(7, "G"), (8, "H"), (9, "I")]:
            cell = ws.cell(
                row=total_row,
                column=col_idx,
                value=f"=SUM({col_letter}2:{col_letter}{last_row})",
            )
            cell.number_format = "#,##0.00 €"
            cell.font = Font(bold=True)
        total_fill = PatternFill(start_color="EEEDFE", end_color="EEEDFE", fill_type="solid")
        for cell in ws[total_row]:
            cell.fill = total_fill

    # Auto-width
    for col in ws.columns:
        try:
            max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
        except Exception:
            continue

    ws.freeze_panes = "A2"
    wb.save(str(output_path))
    logger.info("XLSX registre amortissements généré: %s (%d immos)", output_path.name, len(immos))
    return output_path


# ═══════════════════════════════════════════════════════════════════
# DOTATIONS
# ═══════════════════════════════════════════════════════════════════


def render_dotations(year: int, output_path: Path, format: str, filters: dict) -> Path:
    """Tableau des dotations de l'exercice.

    Filtres : `year` (req), `poste` (all|<poste>).
    """
    from backend.services import amortissement_service

    detail = amortissement_service.get_virtual_detail(year)
    poste = (filters.get("poste") or "all").lower()
    if poste != "all":
        detail.immos = [i for i in detail.immos if (i.poste or "").lower() == poste]
        # Recalcule les totaux post-filtre
        detail.total_brute = float(sum(i.dotation_brute for i in detail.immos))
        detail.total_deductible = float(sum(i.dotation_deductible for i in detail.immos))
        detail.nb_immos_actives = len(detail.immos)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if format == "pdf":
        return _render_dotations_pdf(detail, year, output_path, filters)
    if format == "csv":
        return _render_dotations_csv(detail, year, output_path, filters)
    if format in ("xlsx", "excel"):
        return _render_dotations_xlsx(detail, year, output_path, filters)
    raise ValueError(f"Format non supporté pour dotations : {format}")


def _render_dotations_pdf(detail, year: int, output_path: Path, filters: dict) -> Path:
    """PDF dotations exercice — structure miroir B1 + colonne Origine."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Image,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=landscape(A4),  # paysage : 10 colonnes avec Origine
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )

    styles = getSampleStyleSheet()
    story: list = []

    # Logo
    logo_path = ASSETS_DIR / "logo_lockup_light_400.png"
    if logo_path.exists():
        story.append(Image(str(logo_path), width=55 * mm, height=14 * mm))
        story.append(Spacer(1, 4 * mm))

    # Titre + sous-filtres
    filtre_label = _describe_filters(filters)
    title_text = f"État des amortissements — Exercice {year}"
    if filtre_label:
        title_text += f" · {filtre_label}"
    story.append(Paragraph(title_text, styles["Title"]))
    story.append(Spacer(1, 3 * mm))

    # Tableau registre dotations — 10 colonnes avec Origine
    headers = [
        "Désignation",
        "Origine",
        "Acquis le",
        "Durée",
        "Base",
        "VNC début",
        "Dotation",
        "Q-part",
        "VNC fin",
        "Poste",
    ]
    data: list[list[str]] = [headers]

    if detail.nb_immos_actives == 0:
        data.append(["Aucune immobilisation active sur l'exercice", "", "", "", "", "", "", "", "", ""])
    else:
        for immo in detail.immos:
            designation = immo.designation or ""
            if len(designation) > 35:
                designation = designation[:35] + "…"
            data.append([
                designation,
                _origine_label(immo),
                _format_date_fr(immo.date_acquisition),
                f"{immo.duree} ans",
                _fr_euro(immo.base_amortissable),
                _fr_euro(immo.vnc_debut),
                _fr_euro(immo.dotation_deductible),
                f"{immo.quote_part_pro:.0f} %",
                _fr_euro(immo.vnc_fin),
                immo.poste or "—",
            ])

        # Ligne TOTAL
        data.append([
            "TOTAL", "", "", "", "", "",
            _fr_euro(detail.total_deductible),
            "", "", "",
        ])

    table = Table(
        data,
        repeatRows=1,
        colWidths=[
            48 * mm, 22 * mm, 20 * mm, 14 * mm, 24 * mm,
            24 * mm, 26 * mm, 14 * mm, 24 * mm, 22 * mm,
        ],
    )

    style_cmds: list = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3C3489")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
        ("ALIGN", (4, 1), (8, -1), "RIGHT"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]

    if detail.nb_immos_actives > 0:
        style_cmds += [
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#EEEDFE")),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ]
        # Badge Reprise dans Origine
        for row_idx, immo in enumerate(detail.immos, start=1):
            if _is_reprise(immo):
                style_cmds.append(
                    ("BACKGROUND", (1, row_idx), (1, row_idx), colors.HexColor("#FAEEDA"))
                )
                style_cmds.append(
                    ("TEXTCOLOR", (1, row_idx), (1, row_idx), colors.HexColor("#854F0B"))
                )
                style_cmds.append(
                    ("FONTNAME", (1, row_idx), (1, row_idx), "Helvetica-Bold")
                )

    table.setStyle(TableStyle(style_cmds))
    story.append(table)
    story.append(Spacer(1, 6 * mm))

    # Récapitulatif
    recap_text = (
        f"<b>Dotation brute de l'exercice :</b> {_fr_euro(detail.total_brute)}<br/>"
        f"<b>Dotation déductible :</b> {_fr_euro(detail.total_deductible)}<br/>"
        f"<b>Immobilisations actives :</b> {detail.nb_immos_actives}"
    )
    story.append(Paragraph(recap_text, styles["Normal"]))
    story.append(Spacer(1, 8 * mm))

    # Référence légale
    ref_style = ParagraphStyle(
        "ref",
        parent=styles["Italic"],
        fontSize=8,
        textColor=colors.HexColor("#666666"),
    )
    story.append(Paragraph(
        "Référence : art. 39-1-2° du CGI, PCG art. 214-13 (amortissement linéaire "
        "avec pro rata temporis). Régime BNC — comptabilité de recettes.",
        ref_style,
    ))

    doc.build(story)
    logger.info(
        "PDF dotation amortissements généré: %s (%d immos, %.2f € déductible)",
        output_path.name, detail.nb_immos_actives, detail.total_deductible,
    )
    return output_path


def _render_dotations_csv(detail, year: int, output_path: Path, filters: dict) -> Path:
    """CSV BOM UTF-8, séparateur `;`, virgule décimale, CRLF (Excel FR)."""
    headers = [
        "Désignation",
        "Origine",
        "Acquis le",
        "Mode",
        "Durée",
        "Base",
        "VNC début",
        "Dotation brute",
        "Quote-part (%)",
        "Dotation déductible",
        "VNC fin",
        "Poste",
    ]
    lines: list[str] = [";".join(headers)]

    for immo in detail.immos:
        lines.append(";".join([
            (immo.designation or "").replace(";", ","),
            _origine_label(immo),
            _format_date_fr(immo.date_acquisition),
            immo.mode,
            str(immo.duree),
            _fr_decimal(immo.base_amortissable),
            _fr_decimal(immo.vnc_debut),
            _fr_decimal(immo.dotation_brute),
            f"{immo.quote_part_pro:.0f}",
            _fr_decimal(immo.dotation_deductible),
            _fr_decimal(immo.vnc_fin),
            (immo.poste or "").replace(";", ","),
        ]))

    # Ligne TOTAL
    lines.append(";".join([
        "TOTAL", "", "", "", "", "",
        "",
        _fr_decimal(detail.total_brute),
        "",
        _fr_decimal(detail.total_deductible),
        "", "",
    ]))

    content = "\r\n".join(lines)
    output_path.write_bytes(b"\xef\xbb\xbf" + content.encode("utf-8"))
    logger.info(
        "CSV dotation amortissements généré: %s (%d immos)",
        output_path.name, detail.nb_immos_actives,
    )
    return output_path


def _render_dotations_xlsx(detail, year: int, output_path: Path, filters: dict) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = f"Dotations {year}"[:31]

    headers = [
        "Désignation",
        "Origine",
        "Acquis le",
        "Mode",
        "Durée",
        "Base",
        "VNC début",
        "Dotation brute",
        "Quote-part (%)",
        "Dotation déductible",
        "VNC fin",
        "Poste",
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="3C3489", end_color="3C3489", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    reprise_fill = PatternFill(start_color="FAEEDA", end_color="FAEEDA", fill_type="solid")
    reprise_font = Font(color="854F0B", bold=True)

    for immo in detail.immos:
        row = [
            immo.designation or "",
            _origine_label(immo),
            immo.date_acquisition,
            immo.mode,
            int(immo.duree),
            float(immo.base_amortissable),
            float(immo.vnc_debut),
            float(immo.dotation_brute),
            float(immo.quote_part_pro),
            float(immo.dotation_deductible),
            float(immo.vnc_fin),
            immo.poste or "",
        ]
        ws.append(row)
        if _is_reprise(immo):
            cell = ws.cell(row=ws.max_row, column=2)
            cell.fill = reprise_fill
            cell.font = reprise_font

    # Format EUR colonnes F/G/H/J/K (6/7/8/10/11)
    for col_letter in ("F", "G", "H", "J", "K"):
        for cell in ws[col_letter][1:]:
            cell.number_format = "#,##0.00 €"
            cell.alignment = Alignment(horizontal="right")

    # Ligne TOTAL avec formules SUM
    last_row = ws.max_row
    if last_row >= 2:
        total_row = last_row + 1
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        for col_idx, col_letter in [(8, "H"), (10, "J")]:
            cell = ws.cell(
                row=total_row,
                column=col_idx,
                value=f"=SUM({col_letter}2:{col_letter}{last_row})",
            )
            cell.number_format = "#,##0.00 €"
            cell.font = Font(bold=True)
        total_fill = PatternFill(start_color="EEEDFE", end_color="EEEDFE", fill_type="solid")
        for cell in ws[total_row]:
            cell.fill = total_fill

    for col in ws.columns:
        try:
            max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
        except Exception:
            continue

    ws.freeze_panes = "A2"
    wb.save(str(output_path))
    logger.info(
        "XLSX dotation amortissements généré: %s (%d immos)",
        output_path.name, detail.nb_immos_actives,
    )
    return output_path


# ═══════════════════════════════════════════════════════════════════
# BACKWARD COMPAT (B1 → B3)
# ═══════════════════════════════════════════════════════════════════


def generate_dotation_pdf(year: int, output_path: Path) -> Path:
    """DEPRECATED — wrapper backward-compat B1.

    Délègue à `render_dotations(year, path, "pdf", {"poste": "all"})`.
    Conservé pour ne pas casser les éventuels appels directs (ex. scripts CLI).
    Les chemins de production (OD dotation, export ZIP) consomment désormais
    le template `amortissements_dotations` via `report_service.get_or_generate`.
    """
    return render_dotations(year, output_path, "pdf", {"poste": "all"})
