"""
Rendu des rapports « Compte d'attente sans justificatif » multi-format
(template Rapports V2 `compte_attente_sans_justif`).

Une fonction de rendu unifiée consommée par le template :
  - `render_compte_attente(year, output_path, format, filters)` :
    opérations en compte d'attente / non catégorisées, filtrables par
    mois, scope (justif présent/absent), catégorie/sous-catégorie, source,
    montant et type opération.

Chaque appel dispatche selon le format (`pdf` / `csv` / `xlsx`).
Pattern strict miroir de `amortissement_report_service`.
"""
from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

from backend.core.config import APP_NAME, ASSETS_DIR, MOIS_FR

logger = logging.getLogger(__name__)


# ─── Helpers communs ────────────────────────────────────────────────


def _fr_euro(montant: Optional[float]) -> str:
    """1 234,56 €"""
    if montant is None or montant == 0:
        return "—"
    formatted = f"{abs(montant):,.2f}"
    sign = "-" if montant < 0 else ""
    return f"{sign}{formatted}".replace(",", " ").replace(".", ",") + " €"


def _fr_decimal(montant: Optional[float]) -> str:
    """1234,56 — sans symbole € (CSV/XLSX)."""
    if montant is None:
        return "0,00"
    return f"{montant:.2f}".replace(".", ",")


def _format_date_fr(iso: str) -> str:
    """2024-03-15 → 15/03/2024"""
    if not iso or len(iso) < 10:
        return iso or ""
    return f"{iso[8:10]}/{iso[5:7]}/{iso[0:4]}"


def _describe_filters(filters: dict) -> str:
    """Concatène les filtres actifs pour le sous-titre PDF."""
    parts: list[str] = []
    scope = (filters.get("scope") or "all").lower()
    if filters.get("justificatif_present") is False and scope == "all":
        scope = "sans_justif"
    if scope == "sans_justif":
        parts.append("scope : sans justificatif")

    cats = filters.get("categories") or []
    if cats:
        if len(cats) <= 3:
            parts.append("cat. : " + ", ".join(cats))
        else:
            parts.append(f"cat. : {', '.join(cats[:2])}… (+{len(cats) - 2})")

    subcats = filters.get("subcategories") or []
    if subcats:
        if len(subcats) <= 3:
            parts.append("sous-cat. : " + ", ".join(subcats))
        else:
            parts.append(f"sous-cat. : {', '.join(subcats[:2])}… (+{len(subcats) - 2})")

    source = (filters.get("source") or "all").lower()
    if source and source != "all":
        parts.append(f"source : {source}")

    return " · ".join(parts)


def _resolve_scope(filters: dict) -> dict:
    """Convertit `scope` UI → `justificatif_present` backend.

    Retourne une copie de filters avec `justificatif_present` set selon le scope.
    Si l'utilisateur a déjà fourni `justificatif_present`, on le respecte.
    """
    out = dict(filters)
    if out.get("justificatif_present") is None:
        scope = (out.get("scope") or "").lower()
        if scope == "sans_justif":
            out["justificatif_present"] = False
    return out


def _safe_float(val) -> float:
    try:
        if val is None:
            return 0.0
        return float(val)
    except (TypeError, ValueError):
        return 0.0


# ─── Collecte ───────────────────────────────────────────────────────


def _collect_attente_ops(year: int, month: Optional[int]) -> list[dict]:
    """Charge les opérations en compte d'attente / non catégorisées, format natif.

    Filtre :
      - `compte_attente == True` OU
      - catégorie vide / "Autres" / "Non catégorisé" / "?"

    Retourne des dicts au format natif (`Date`, `Libellé`, `Catégorie`,
    `Sous-catégorie`, `Débit`, `Crédit`, `Lien justificatif`, `Justificatif`,
    `Commentaire`, `alertes`, `alertes_resolues`, `compte_attente`,
    `alerte_note`, `source`, `ventilation`, `_filename`, `_index`).

    Les ventilations sont conservées en bloc pour être éclatées plus tard.
    """
    from backend.services import operation_service

    files = operation_service.list_operation_files()
    files = [f for f in files if f.get("year") == year]
    if month:
        files = [f for f in files if f.get("month") == month]

    out: list[dict] = []
    for f in files:
        try:
            ops = operation_service.load_operations(f["filename"])
        except Exception as e:
            logger.warning("Impossible de charger %s: %s", f["filename"], e)
            continue

        for idx, op in enumerate(ops):
            cat = (op.get("Catégorie") or "").strip()
            is_attente = bool(op.get("compte_attente"))
            is_empty_cat = cat in ("", "Autres", "Non catégorisé", "?")

            if not (is_attente or is_empty_cat):
                continue

            enriched = dict(op)
            enriched["_filename"] = f["filename"]
            enriched["_index"] = idx
            out.append(enriched)

    out.sort(key=lambda o: (o.get("Date") or "", o.get("_filename") or "", o.get("_index") or 0))
    return out


def _explode_with_alertes(operations: list[dict]) -> list[dict]:
    """Éclate les ventilations en sous-lignes tout en propageant les alertes du parent.

    Pattern miroir de `report_service._explode_ventilations` mais propage en plus :
    `alertes`, `alertes_resolues`, `compte_attente`, `alerte_note`, `source`.
    Les ops non ventilées sont passées inchangées.
    """
    out: list[dict] = []
    for op in operations:
        vlines = op.get("ventilation") or []
        if not vlines:
            out.append(op)
            continue
        parent_libelle = op.get("Libellé", "") or ""
        parent_debit_positive = _safe_float(op.get("Débit", 0)) > 0
        parent_credit_positive = _safe_float(op.get("Crédit", 0)) > 0
        parent_alertes = op.get("alertes", []) or []
        parent_alertes_resolues = op.get("alertes_resolues", []) or []
        parent_compte_attente = bool(op.get("compte_attente"))
        parent_alerte_note = op.get("alerte_note") or ""
        parent_source = op.get("source")
        n = len(vlines)
        for i, vl in enumerate(vlines):
            montant = _safe_float(vl.get("montant", 0))
            vl_justif = (vl.get("justificatif") or "").strip()
            out.append({
                "Date": op.get("Date", ""),
                "Libellé": f"{parent_libelle} [V{i+1}/{n}]",
                "Catégorie": vl.get("categorie", "") or "",
                "Sous-catégorie": vl.get("sous_categorie", "") or "",
                "Débit": montant if parent_debit_positive else 0,
                "Crédit": montant if parent_credit_positive else 0,
                "Justificatif": bool(vl_justif),
                "Lien justificatif": f"traites/{vl_justif}" if vl_justif else "",
                "Commentaire": op.get("Commentaire", "") or "",
                "Important": op.get("Important", False),
                "alertes": list(parent_alertes),
                "alertes_resolues": list(parent_alertes_resolues),
                "compte_attente": parent_compte_attente,
                "alerte_note": parent_alerte_note,
                "source": parent_source,
                "_filename": op.get("_filename"),
                "_index": op.get("_index"),
                "_ventilation_index": i,
            })
    return out


def _alertes_label(op: dict) -> str:
    alertes = op.get("alertes") or []
    if alertes:
        return ", ".join(alertes)
    if op.get("compte_attente"):
        return "a_categoriser"
    if not (op.get("Catégorie") or "").strip():
        return "a_categoriser"
    return ""


# ─── Entrée principale ─────────────────────────────────────────────


def render_compte_attente(year: int, output_path: Path, format: str, filters: dict) -> Path:
    """Rapport « Compte d'attente sans justificatif » filtrable.

    Filtres supportés : `year` (req), `month`, `scope` (all|sans_justif),
    `justificatif_present` (bool), `categories`, `subcategories`, `source`,
    `min_amount`, `max_amount`, `type` (debit|credit).
    """
    from backend.services import report_service

    resolved = _resolve_scope(filters)
    month = resolved.get("month")
    if month is not None:
        try:
            month = int(month)
        except (TypeError, ValueError):
            month = None

    raw_ops = _collect_attente_ops(year, month)
    exploded = _explode_with_alertes(raw_ops)
    filtered = report_service._apply_filters(exploded, resolved)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if format == "pdf":
        return _render_pdf(filtered, year, month, output_path, resolved)
    if format == "csv":
        return _render_csv(filtered, year, month, output_path, resolved)
    if format in ("xlsx", "excel"):
        return _render_xlsx(filtered, year, month, output_path, resolved)
    raise ValueError(f"Format non supporté pour compte d'attente : {format}")


# ─── PDF ───────────────────────────────────────────────────────────


def _render_pdf(operations: list[dict], year: int, month: Optional[int],
                output_path: Path, filters: dict) -> Path:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import (
        BaseDocTemplate,
        Frame,
        Image,
        PageTemplate,
        Paragraph,
        Spacer,
        Table,
        TableStyle,
    )

    page_w, page_h = landscape(A4)
    margin_lr = 1.5 * cm
    margin_tb = 1.5 * cm

    period_label = f"{MOIS_FR[month - 1].capitalize()} {year}" if month else f"Année {year}"

    def _footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setStrokeColor(colors.HexColor("#CCCCCC"))
        y_line = margin_tb - 8 * mm
        canvas.line(margin_lr, y_line, page_w - margin_lr, y_line)
        canvas.drawString(margin_lr, y_line - 4 * mm, f"Page {doc.page}")
        gen_date = datetime.now().strftime("%d/%m/%Y %H:%M")
        right_text = f"{APP_NAME} — Compte d'attente — {period_label} — {gen_date}"
        canvas.drawRightString(page_w - margin_lr, y_line - 4 * mm, right_text)
        canvas.restoreState()

    doc = BaseDocTemplate(
        str(output_path),
        pagesize=landscape(A4),
        leftMargin=margin_lr,
        rightMargin=margin_lr,
        topMargin=margin_tb,
        bottomMargin=margin_tb + 5 * mm,
    )
    usable_w = page_w - 2 * margin_lr
    frame = Frame(margin_lr, margin_tb, usable_w,
                  page_h - 2 * margin_tb - 5 * mm, id="main")
    doc.addPageTemplates([PageTemplate(id="main", frames=[frame], onPage=_footer)])

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("CATitle", parent=styles["Title"], fontSize=14,
                                  spaceAfter=4, textColor=colors.HexColor("#3C3489"))
    sub_style = ParagraphStyle("CASub", parent=styles["Normal"], fontSize=9,
                                textColor=colors.HexColor("#666666"), spaceAfter=8)
    cell_style = ParagraphStyle("CACell", parent=styles["Normal"], fontSize=7, leading=9)
    cell_right = ParagraphStyle("CACellR", parent=cell_style, alignment=TA_RIGHT)
    just_yes_style = ParagraphStyle("CAJustYes", parent=cell_style,
                                     textColor=colors.HexColor("#16a34a"), alignment=TA_CENTER)
    just_no_style = ParagraphStyle("CAJustNo", parent=cell_style,
                                    textColor=colors.HexColor("#aaaaaa"), alignment=TA_CENTER)
    empty_style = ParagraphStyle("CAEmpty", parent=styles["Normal"], fontSize=11,
                                  textColor=colors.HexColor("#888888"), alignment=TA_CENTER)

    story: list = []

    # Logo
    logo_path = ASSETS_DIR / "logo_lockup_light_400.png"
    if logo_path.exists():
        try:
            logo = Image(str(logo_path), width=120, height=40)
            logo.hAlign = "LEFT"
            story.append(logo)
            story.append(Spacer(1, 6))
        except Exception:
            pass

    # Titre
    scope = (filters.get("scope") or "").lower()
    if filters.get("justificatif_present") is False and scope != "sans_justif":
        scope = "sans_justif"
    title_text = "Compte d'attente — sans justificatif" if scope == "sans_justif" else "Compte d'attente"
    story.append(Paragraph(f"{title_text} — {period_label}", title_style))

    filter_desc = _describe_filters(filters)
    sub_text = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} | {len(operations)} opération(s)"
    if filter_desc:
        sub_text += f" | {filter_desc}"
    story.append(Paragraph(sub_text, sub_style))
    story.append(Spacer(1, 8))

    if not operations:
        story.append(Spacer(1, 40))
        story.append(Paragraph("Aucune opération ne correspond aux filtres actifs", empty_style))
        story.append(Spacer(1, 20))
    else:
        # Tableau
        headers = ["Date", "Libellé", "Catégorie", "Sous-cat.", "Débit", "Crédit",
                   "Type alerte", "Just.", "Commentaire"]
        data: list = [headers]
        # Largeurs colonnes (paysage A4 ≈ 26 cm utiles)
        col_widths = [50, 200, 70, 60, 60, 60, 70, 30, 90]

        for op in operations:
            debit = _safe_float(op.get("Débit", 0))
            credit = _safe_float(op.get("Crédit", 0))
            lien = (op.get("Lien justificatif") or "").strip()
            has_just = bool(lien)
            comment = str(op.get("Commentaire") or op.get("alerte_note") or "")

            data.append([
                _format_date_fr(op.get("Date", "")),
                Paragraph(str(op.get("Libellé", ""))[:120], cell_style),
                str(op.get("Catégorie", "")),
                str(op.get("Sous-catégorie", "")),
                Paragraph(_fr_euro(debit) if debit else "", cell_right),
                Paragraph(_fr_euro(credit) if credit else "", cell_right),
                Paragraph(_alertes_label(op), cell_style),
                Paragraph("☑" if has_just else "☐",
                          just_yes_style if has_just else just_no_style),
                Paragraph(comment[:60], cell_style) if comment else "",
            ])

        # Ligne TOTAL
        total_debit = sum(_safe_float(o.get("Débit", 0)) for o in operations)
        total_credit = sum(_safe_float(o.get("Crédit", 0)) for o in operations)
        nb_just = sum(1 for o in operations if (o.get("Lien justificatif") or "").strip())
        data.append([
            "", "TOTAL", "", "",
            _fr_euro(total_debit), _fr_euro(total_credit),
            "", f"{nb_just}/{len(operations)}", "",
        ])

        n_rows = len(data)
        style_cmds: list = [
            # Header
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3C3489")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            # Body
            ("FONTSIZE", (0, 1), (-1, -1), 7),
            ("ALIGN", (4, 1), (5, -1), "RIGHT"),
            ("ALIGN", (7, 1), (7, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#DDDDDD")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F9F9F9")]),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            # Ligne TOTAL violette
            ("BACKGROUND", (0, n_rows - 1), (-1, n_rows - 1), colors.HexColor("#EEEDFE")),
            ("FONTNAME", (0, n_rows - 1), (-1, n_rows - 1), "Helvetica-Bold"),
            ("FONTSIZE", (0, n_rows - 1), (-1, n_rows - 1), 8),
            ("LINEABOVE", (0, n_rows - 1), (-1, n_rows - 1), 1.5, colors.HexColor("#3C3489")),
        ]

        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle(style_cmds))
        story.append(table)
        story.append(Spacer(1, 8))

        # Récapitulatif par catégorie
        cat_stats: dict[str, dict] = {}
        for op in operations:
            cat = (op.get("Catégorie") or "Non catégorisé") or "Non catégorisé"
            stat = cat_stats.setdefault(cat, {"debit": 0.0, "credit": 0.0, "count": 0, "no_just": 0})
            stat["debit"] += _safe_float(op.get("Débit", 0))
            stat["credit"] += _safe_float(op.get("Crédit", 0))
            stat["count"] += 1
            if not (op.get("Lien justificatif") or "").strip():
                stat["no_just"] += 1

        if cat_stats:
            recap_headers = ["Catégorie", "Nb ops", "Sans justif", "Total Débits", "Total Crédits"]
            recap_data: list = [recap_headers]
            for cat, s in sorted(cat_stats.items()):
                recap_data.append([
                    cat,
                    str(s["count"]),
                    str(s["no_just"]),
                    _fr_euro(s["debit"]),
                    _fr_euro(s["credit"]),
                ])
            recap_table = Table(recap_data, colWidths=[200, 80, 80, 100, 100])
            recap_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3C3489")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("ALIGN", (1, 1), (4, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
            ]))
            story.append(Paragraph("<b>Récapitulatif par catégorie</b>",
                                    ParagraphStyle("recap", parent=styles["Normal"], fontSize=10,
                                                   textColor=colors.HexColor("#3C3489"))))
            story.append(Spacer(1, 4))
            story.append(recap_table)

    doc.build(story)
    logger.info(
        "PDF compte d'attente généré: %s (%d ops)",
        output_path.name, len(operations),
    )
    return output_path


# ─── CSV ───────────────────────────────────────────────────────────


def _render_csv(operations: list[dict], year: int, month: Optional[int],
                output_path: Path, filters: dict) -> Path:
    """CSV BOM UTF-8, séparateur `;`, virgule décimale, CRLF (Excel FR)."""
    headers = ["Date", "Libellé", "Catégorie", "Sous-catégorie", "Débit", "Crédit",
               "Type alerte", "Justificatif", "Commentaire"]
    lines: list[str] = [";".join(headers)]

    for op in operations:
        debit = _safe_float(op.get("Débit", 0))
        credit = _safe_float(op.get("Crédit", 0))
        libelle = str(op.get("Libellé", "") or "").replace(";", ",")
        comment = str(op.get("Commentaire") or op.get("alerte_note") or "").replace(";", ",")
        lien = (op.get("Lien justificatif") or "").strip()
        just_str = "1" if lien else "0"
        lines.append(";".join([
            op.get("Date", ""),
            libelle,
            str(op.get("Catégorie", "") or "").replace(";", ","),
            str(op.get("Sous-catégorie", "") or "").replace(";", ","),
            _fr_decimal(debit) if debit else "",
            _fr_decimal(credit) if credit else "",
            _alertes_label(op),
            just_str,
            comment,
        ]))

    # Ligne TOTAL
    total_debit = sum(_safe_float(o.get("Débit", 0)) for o in operations)
    total_credit = sum(_safe_float(o.get("Crédit", 0)) for o in operations)
    nb_just = sum(1 for o in operations if (o.get("Lien justificatif") or "").strip())
    lines.append("")
    lines.append(";".join([
        "", "TOTAL", "", "",
        _fr_decimal(total_debit),
        _fr_decimal(total_credit),
        "", f"{nb_just}/{len(operations)}", "",
    ]))

    content = "\r\n".join(lines) + "\r\n"
    output_path.write_bytes(b"\xef\xbb\xbf" + content.encode("utf-8"))
    logger.info(
        "CSV compte d'attente généré: %s (%d ops)",
        output_path.name, len(operations),
    )
    return output_path


# ─── XLSX ──────────────────────────────────────────────────────────


def _render_xlsx(operations: list[dict], year: int, month: Optional[int],
                 output_path: Path, filters: dict) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    period_short = f"{year}-{month:02d}" if month else str(year)
    ws.title = f"CA {period_short}"[:31]

    headers = ["Date", "Libellé", "Catégorie", "Sous-catégorie", "Débit", "Crédit",
               "Type alerte", "Justificatif", "Commentaire"]
    ws.append(headers)

    header_fill = PatternFill(start_color="3C3489", end_color="3C3489", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for op in operations:
        debit = _safe_float(op.get("Débit", 0))
        credit = _safe_float(op.get("Crédit", 0))
        lien = (op.get("Lien justificatif") or "").strip()
        ws.append([
            op.get("Date", ""),
            op.get("Libellé", "") or "",
            op.get("Catégorie", "") or "",
            op.get("Sous-catégorie", "") or "",
            debit,
            credit,
            _alertes_label(op),
            "Oui" if lien else "Non",
            op.get("Commentaire") or op.get("alerte_note") or "",
        ])

    # Format EUR sur colonnes E/F (5/6)
    for col_letter in ("E", "F"):
        for cell in ws[col_letter][1:]:
            cell.number_format = "#,##0.00 €"
            cell.alignment = Alignment(horizontal="right")

    # Ligne TOTAL avec formules SUM
    last_row = ws.max_row
    if last_row >= 2:
        total_row = last_row + 1
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=2, value=f"{len(operations)} op(s)")
        for col_idx, col_letter in [(5, "E"), (6, "F")]:
            cell = ws.cell(
                row=total_row,
                column=col_idx,
                value=f"=SUM({col_letter}2:{col_letter}{last_row})",
            )
            cell.number_format = "#,##0.00 €"
            cell.font = Font(bold=True)
        nb_just = sum(1 for o in operations if (o.get("Lien justificatif") or "").strip())
        ws.cell(row=total_row, column=8, value=f"{nb_just}/{len(operations)}").font = Font(bold=True)
        total_fill = PatternFill(start_color="EEEDFE", end_color="EEEDFE", fill_type="solid")
        for cell in ws[total_row]:
            cell.fill = total_fill

    # Auto-width
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            try:
                length = len(str(cell.value)) if cell.value is not None else 0
            except Exception:
                length = 0
            if length > max_len:
                max_len = length
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 50)

    ws.freeze_panes = "A2"

    # Onglet récap par catégorie
    if operations:
        ws2 = wb.create_sheet("Récap par catégorie")
        ws2.append(["Catégorie", "Nb ops", "Sans justif", "Total Débits", "Total Crédits"])
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        cat_stats: dict[str, dict] = {}
        for op in operations:
            cat = (op.get("Catégorie") or "Non catégorisé") or "Non catégorisé"
            stat = cat_stats.setdefault(cat, {"debit": 0.0, "credit": 0.0, "count": 0, "no_just": 0})
            stat["debit"] += _safe_float(op.get("Débit", 0))
            stat["credit"] += _safe_float(op.get("Crédit", 0))
            stat["count"] += 1
            if not (op.get("Lien justificatif") or "").strip():
                stat["no_just"] += 1

        for cat, s in sorted(cat_stats.items()):
            ws2.append([cat, s["count"], s["no_just"], s["debit"], s["credit"]])

        for col_letter in ("D", "E"):
            for cell in ws2[col_letter][1:]:
                cell.number_format = "#,##0.00 €"
                cell.alignment = Alignment(horizontal="right")

        for col_idx in range(1, 6):
            col_letter = get_column_letter(col_idx)
            max_len = max((len(str(c.value)) if c.value is not None else 0) for c in ws2[col_letter])
            ws2.column_dimensions[col_letter].width = min(max_len + 2, 40)
        ws2.freeze_panes = "A2"

    wb.save(str(output_path))
    logger.info(
        "XLSX compte d'attente généré: %s (%d ops)",
        output_path.name, len(operations),
    )
    return output_path
