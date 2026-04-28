"""
Service pour la génération de rapports V2.
Index JSON, templates, format EUR, déduplication, réconciliation.

Templates « custom renderer » (Prompt B3) :
  Un template peut porter les champs optionnels `category`, `formats`,
  `filters_schema`, `renderer` (callable), `dedup_key_fn`, `title_builder`.
  Quand `renderer` est présent, `generate_report()` délègue à `get_or_generate()`
  qui dispatch sur le moteur custom au lieu de la pipeline ops bancaires standard.
  Les templates amortissements consomment ce mécanisme pour partager un seul
  point de génération entre l'UI Rapports V2, l'OD dotation et l'export ZIP.
"""
from __future__ import annotations

import csv
import hashlib
import io
import json
import logging
import os
import re
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Optional

import shutil

from backend.core.config import (
    RAPPORTS_DIR, REPORTS_DIR, REPORTS_INDEX, REPORTS_ARCHIVES_DIR, ASSETS_DIR,
    MOIS_FR, ensure_directories,
)
from backend.services.operation_service import list_operation_files, load_operations

logger = logging.getLogger(__name__)

# ─── Templates prédéfinis ───
#
# Note: les champs `renderer`, `dedup_key_fn`, `title_builder` sont des callables
# Python non-sérialisables. `get_templates()` les filtre avant exposition API.

REPORT_TEMPLATES: list[dict] = [
    {
        "id": "bnc_annuel",
        "label": "Récapitulatif annuel BNC",
        "description": "Toutes les opérations de l'année — recettes et dépenses",
        "icon": "FileText",
        "format": "pdf",
        "filters": {"type": "all"},
    },
    {
        "id": "ventilation_charges",
        "label": "Ventilation des charges",
        "description": "Dépenses par catégorie et sous-catégorie",
        "icon": "PieChart",
        "format": "excel",
        "filters": {"type": "debit"},
    },
    {
        "id": "recapitulatif_social",
        "label": "Récapitulatif social",
        "description": "Charges URSSAF, CARMF, ODM sur l'année",
        "icon": "Shield",
        "format": "pdf",
        "filters": {"categories": ["Charges sociales"], "type": "debit"},
    },
    {
        "id": "amortissements_registre",
        "label": "Registre des immobilisations",
        "description": "État complet du registre (actives, amorties, sorties)",
        "icon": "Package",
        "category": "Amortissements",
        "format": "pdf",
        "formats": ["pdf", "csv", "xlsx"],
        "filters": {"statut": "all", "poste": "all"},
        "filters_schema": [
            {"key": "year", "type": "int", "required": True},
            {"key": "statut", "type": "select",
             "options": ["all", "en_cours", "amorti", "sorti"], "default": "all"},
            {"key": "poste", "type": "select", "options": "dynamic:postes", "default": "all"},
        ],
        # renderer / dedup_key_fn / title_builder injectés via _ensure_amortissements_renderers
    },
    {
        "id": "amortissements_dotations",
        "label": "Tableau des dotations",
        "description": "Dotations de l'exercice par immobilisation",
        "icon": "TrendingDown",
        "category": "Amortissements",
        "format": "pdf",
        "formats": ["pdf", "csv", "xlsx"],
        "filters": {"poste": "all"},
        "filters_schema": [
            {"key": "year", "type": "int", "required": True},
            {"key": "poste", "type": "select", "options": "dynamic:postes", "default": "all"},
        ],
    },
    {
        "id": "compte_attente_sans_justif",
        "label": "Compte d'attente — sans justificatif",
        "description": "Opérations en attente ou non catégorisées, filtrables par mois/cat/sous-cat/source",
        "icon": "AlertTriangle",
        "category": "Compte d'attente",
        "format": "pdf",
        "formats": ["pdf", "csv", "xlsx"],
        "filters": {"scope": "sans_justif"},
        "filters_schema": [
            {"key": "year", "type": "int", "required": True},
            {"key": "month", "type": "int", "required": False},
            {"key": "scope", "type": "select",
             "options": ["all", "sans_justif"], "default": "sans_justif"},
            {"key": "categories", "type": "multi-select", "options": "dynamic:categories"},
            {"key": "subcategories", "type": "multi-select", "options": "dynamic:subcategories"},
            {"key": "source", "type": "select",
             "options": ["all", "bancaire", "note_de_frais"], "default": "all"},
        ],
        # renderer / dedup_key_fn / title_builder injectés via _ensure_compte_attente_renderers
    },
]

# Champs sérialisables pour exposition API frontend
_TEMPLATE_PUBLIC_KEYS = {
    "id", "label", "description", "icon", "format", "filters",
    "category", "formats", "filters_schema",
}


def _amort_registre_dedup_key(filters: dict) -> str:
    year = filters.get("year") or "all"
    statut = filters.get("statut", "all") or "all"
    poste = filters.get("poste", "all") or "all"
    return f"amort_registre_{year}_{statut}_{poste}"


def _amort_registre_title(filters: dict) -> str:
    year = filters.get("year") or "—"
    base = f"Registre immobilisations {year}"
    statut = filters.get("statut", "all")
    poste = filters.get("poste", "all")
    if statut and statut != "all":
        base += f" · {statut}"
    if poste and poste != "all":
        base += f" · poste {poste}"
    return base


def _amort_dotations_dedup_key(filters: dict) -> str:
    year = filters.get("year") or "all"
    poste = filters.get("poste", "all") or "all"
    return f"amort_dotations_{year}_{poste}"


def _amort_dotations_title(filters: dict) -> str:
    year = filters.get("year") or "—"
    base = f"Tableau dotations {year}"
    poste = filters.get("poste", "all")
    if poste and poste != "all":
        base += f" · poste {poste}"
    return base


def _ensure_amortissements_renderers() -> None:
    """Injecte renderer/dedup_key_fn/title_builder sur les templates amortissements.

    Lazy pour éviter le cycle d'imports `report_service ↔ amortissement_report_service`.
    Idempotent.
    """
    needs_injection = any(
        t["id"] in ("amortissements_registre", "amortissements_dotations")
        and "renderer" not in t
        for t in REPORT_TEMPLATES
    )
    if not needs_injection:
        return
    from backend.services import amortissement_report_service
    for t in REPORT_TEMPLATES:
        if t["id"] == "amortissements_registre":
            t["renderer"] = amortissement_report_service.render_registre
            t["dedup_key_fn"] = _amort_registre_dedup_key
            t["title_builder"] = _amort_registre_title
        elif t["id"] == "amortissements_dotations":
            t["renderer"] = amortissement_report_service.render_dotations
            t["dedup_key_fn"] = _amort_dotations_dedup_key
            t["title_builder"] = _amort_dotations_title


def _compte_attente_dedup_key(filters: dict) -> str:
    year = filters.get("year") or "all"
    month = filters.get("month") or "all"
    scope = (filters.get("scope") or "all").lower() or "all"
    if filters.get("justificatif_present") is False and scope == "all":
        scope = "sans_justif"
    cats = filters.get("categories") or []
    cats_part = "_".join(sorted(c.lower() for c in cats)) if cats else "all"
    subcats = filters.get("subcategories") or []
    subcats_part = "_".join(sorted(s.lower() for s in subcats)) if subcats else "all"
    source = (filters.get("source") or "all").lower() or "all"
    return f"compte_attente_{year}_{month}_{scope}_{cats_part}_{subcats_part}_{source}"


def _compte_attente_title(filters: dict) -> str:
    year = filters.get("year") or "—"
    month = filters.get("month")
    scope = (filters.get("scope") or "all").lower()
    if filters.get("justificatif_present") is False and scope == "all":
        scope = "sans_justif"

    if month and 1 <= int(month) <= 12:
        period = f"{MOIS_FR[int(month) - 1].capitalize()} {year}"
    else:
        period = str(year)

    base = "Compte d'attente"
    if scope == "sans_justif":
        base = "Compte d'attente — sans justificatif"

    cats = filters.get("categories") or []
    if cats and 1 <= len(cats) <= 3:
        base += f" · {', '.join(cats)}"
    elif cats and len(cats) > 3:
        base += f" · {', '.join(cats[:2])}… (+{len(cats) - 2})"
    return f"{base} — {period}"


def _ensure_compte_attente_renderers() -> None:
    """Injecte renderer/dedup_key_fn/title_builder sur le template compte d'attente.

    Lazy pour éviter le cycle d'imports
    `report_service ↔ compte_attente_report_service`.
    Idempotent.
    """
    needs_injection = any(
        t["id"] == "compte_attente_sans_justif" and "renderer" not in t
        for t in REPORT_TEMPLATES
    )
    if not needs_injection:
        return
    from backend.services import compte_attente_report_service
    for t in REPORT_TEMPLATES:
        if t["id"] == "compte_attente_sans_justif":
            t["renderer"] = compte_attente_report_service.render_compte_attente
            t["dedup_key_fn"] = _compte_attente_dedup_key
            t["title_builder"] = _compte_attente_title


def get_templates() -> list[dict]:
    """Retourne les templates avec uniquement les champs sérialisables.

    Les callables (`renderer`, `dedup_key_fn`, `title_builder`) sont strippés
    pour permettre l'exposition JSON via `/api/reports/templates`.
    """
    return [
        {k: v for k, v in t.items() if k in _TEMPLATE_PUBLIC_KEYS}
        for t in REPORT_TEMPLATES
    ]


def _get_template_by_id(template_id: str) -> Optional[dict]:
    _ensure_amortissements_renderers()
    _ensure_compte_attente_renderers()
    for t in REPORT_TEMPLATES:
        if t["id"] == template_id:
            return t
    return None


# ─── Index JSON ───

def _load_index() -> dict:
    if REPORTS_INDEX.exists():
        try:
            with open(REPORTS_INDEX, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            logger.error(f"Erreur chargement index rapports: {e}")
    return {"version": 1, "reports": []}


def _save_index(index: dict) -> None:
    ensure_directories()
    with open(REPORTS_INDEX, "w", encoding="utf-8") as f:
        json.dump(index, f, ensure_ascii=False, indent=2, default=str)


def reconcile_index() -> None:
    """Réconcilie l'index avec le filesystem."""
    ensure_directories()
    index = _load_index()
    reports = index.get("reports", [])

    # Map existant
    existing_files: set[str] = set()
    for d in [RAPPORTS_DIR, REPORTS_DIR]:
        if d.exists():
            for f in d.iterdir():
                if f.suffix in (".pdf", ".csv", ".xlsx"):
                    existing_files.add(f.name)

    # Supprimer les entrées sans fichier
    before = len(reports)
    reports = [r for r in reports if r["filename"] in existing_files]
    removed = before - len(reports)

    # Indexer les fichiers non référencés
    indexed_files = {r["filename"] for r in reports}
    added = 0
    for fname in existing_files:
        if fname not in indexed_files:
            path = _find_report_path(fname)
            if path:
                stat = path.stat()
                ext = path.suffix[1:].lower()
                fmt = "excel" if ext == "xlsx" else ext
                reports.append({
                    "filename": fname,
                    "title": _clean_filename_title(fname),
                    "description": None,
                    "format": fmt,
                    "generated_at": datetime.fromtimestamp(stat.st_mtime).isoformat(),
                    "filters": {},
                    "template_id": None,
                    "nb_operations": 0,
                    "total_debit": 0,
                    "total_credit": 0,
                    "file_size": stat.st_size,
                    "file_size_human": _format_size(stat.st_size),
                    "year": None,
                    "quarter": None,
                    "month": None,
                })
                added += 1

    index["reports"] = reports
    _save_index(index)
    if removed or added:
        logger.info(f"Reports index reconciled: {removed} removed, {added} added")


def _clean_filename_title(filename: str) -> str:
    stem = Path(filename).stem
    stem = re.sub(r"_\d{8}_\d{6}$", "", stem)  # remove timestamp suffix
    stem = re.sub(r"_[a-f0-9]{8}$", "", stem)   # remove hash suffix
    return stem.replace("_", " ").replace("rapport ", "").title()


# ─── Gallery ───

def get_all_reports() -> list[dict]:
    index = _load_index()
    reports = index.get("reports", [])
    reports.sort(key=lambda r: r.get("generated_at", ""), reverse=True)
    return reports


def get_gallery() -> dict:
    reports = get_all_reports()
    years = sorted(set(r.get("year") for r in reports if r.get("year")), reverse=True)
    return {
        "reports": reports,
        "available_years": years,
        "total_count": len(reports),
    }


# ─── CRUD ───

def get_report_path(filename: str) -> Optional[Path]:
    return _find_report_path(filename)


def _find_report_path(filename: str) -> Optional[Path]:
    for d in [RAPPORTS_DIR, REPORTS_DIR]:
        path = d / filename
        if path.exists():
            return path
    return None


def delete_report(filename: str) -> bool:
    path = _find_report_path(filename)
    if path:
        path.unlink()
    # Remove from index
    index = _load_index()
    index["reports"] = [r for r in index["reports"] if r["filename"] != filename]
    _save_index(index)
    return path is not None


def update_report_metadata(filename: str, updates: dict) -> Optional[dict]:
    index = _load_index()
    for r in index["reports"]:
        if r["filename"] == filename:
            if updates.get("title") is not None:
                r["title"] = updates["title"]
            if updates.get("description") is not None:
                r["description"] = updates["description"]
            _save_index(index)
            return r
    return None


# ─── Generation ───

def _slugify(text: str, max_len: int = 40) -> str:
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^\w\s-]", "", text.lower())
    text = re.sub(r"[\s-]+", "_", text).strip("_")
    return text[:max_len]


def _format_eur(amount: float) -> str:
    if amount == 0:
        return "—"
    formatted = f"{abs(amount):,.2f}".replace(",", " ").replace(".", ",")
    sign = "-" if amount < 0 else ""
    return f"{sign}{formatted} €"


def _generate_title(filters: dict, template_id: Optional[str] = None) -> str:
    # Period
    year = filters.get("year")
    quarter = filters.get("quarter")
    month = filters.get("month")
    date_from = filters.get("date_from")
    date_to = filters.get("date_to")

    if year and month and 1 <= month <= 12:
        period = f"{MOIS_FR[month - 1].capitalize()} {year}"
    elif year and quarter:
        period = f"T{quarter} {year}"
    elif year:
        period = str(year)
    elif date_from and date_to:
        period = f"{date_from} au {date_to}"
    else:
        period = "Toutes périodes"

    # Template
    if template_id:
        for t in REPORT_TEMPLATES:
            if t["id"] == template_id:
                return f"{t['label']} — {period}"

    # Categories
    cats = filters.get("categories")
    if cats and len(cats) == 1:
        return f"{cats[0]} — {period}"
    elif cats and 2 <= len(cats) <= 4:
        return f"{', '.join(cats)} — {period}"
    elif cats and len(cats) > 4:
        return f"{', '.join(cats[:3])}… (+{len(cats) - 3}) — {period}"
    else:
        return f"Toutes catégories — {period}"


def _dedup_key(filters: dict, fmt: str) -> tuple:
    return (
        filters.get("year"),
        filters.get("quarter"),
        filters.get("month"),
        filters.get("date_from"),
        filters.get("date_to"),
        tuple(sorted(filters.get("categories") or [])),
        fmt,
    )


def _archive_report(old_path: Path, metadata: dict) -> None:
    """Archive un ancien rapport dans data/reports/archives/ au lieu de le supprimer."""
    REPORTS_ARCHIVES_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    archived_name = f"{old_path.stem}_archived_{timestamp}{old_path.suffix}"
    archive_dest = REPORTS_ARCHIVES_DIR / archived_name
    shutil.move(str(old_path), str(archive_dest))
    logger.info("Rapport archivé: %s → %s", old_path.name, archived_name)

    # Sauvegarder les metadata de la version archivée
    archive_meta_path = REPORTS_ARCHIVES_DIR / "archives_index.json"
    archive_index: list[dict] = []
    if archive_meta_path.exists():
        try:
            archive_index = json.loads(archive_meta_path.read_text(encoding="utf-8"))
        except Exception:
            pass
    archive_index.append({
        "original_filename": old_path.name,
        "archived_filename": archived_name,
        "archived_at": datetime.now().isoformat(),
        "generated_at": metadata.get("generated_at", ""),
        "title": metadata.get("title", ""),
        "format": metadata.get("format", ""),
        "filters": metadata.get("filters", {}),
    })
    archive_meta_path.write_text(json.dumps(archive_index, ensure_ascii=False, indent=2), encoding="utf-8")


def get_or_generate(
    template_id: str,
    filters: dict,
    format: str,
    title: Optional[str] = None,
    description: Optional[str] = None,
) -> dict:
    """Retourne le rapport existant matchant la `dedup_key` + format, sinon le génère.

    Élimine la duplication entre les 3 sources historiques (UI Rapports / OD dotation
    / export ZIP) — cf. Prompt B3.

    Pour les templates « custom renderer » (champ `renderer` présent), utilise
    `dedup_key_fn(filters)` pour calculer une clé stable et `renderer(year, path,
    format, filters)` pour générer. Pour les templates standards (BNC annuel,
    Ventilation charges, etc.), délègue à `generate_report()` (legacy path).

    Retourne le dict metadata du rapport (mêmes champs que `generate_report`),
    avec `from_cache: True` quand un fichier existant a été réutilisé.
    """
    ensure_directories()
    template = _get_template_by_id(template_id)
    if not template:
        raise ValueError(f"Template inconnu : {template_id}")

    # Templates standards (sans renderer) → legacy path
    if not template.get("renderer"):
        return generate_report({
            "template_id": template_id,
            "filters": filters,
            "format": format,
            "title": title,
            "description": description,
        })

    fmt = format
    dedup_key_fn = template["dedup_key_fn"]
    title_builder = template.get("title_builder", lambda f: template["label"])
    title_str = title or title_builder(filters)
    dedup_key_str = dedup_key_fn(filters)

    # 1. Recherche d'un rapport existant matchant `dedup_key + format`
    index = _load_index()
    for existing in index["reports"]:
        if (
            existing.get("dedup_key") == dedup_key_str
            and existing.get("format") == fmt
        ):
            existing_path = _find_report_path(existing["filename"])
            if existing_path and existing_path.exists():
                logger.info(
                    "get_or_generate cache hit: %s (%s)",
                    existing["filename"], dedup_key_str,
                )
                return {**existing, "replaced": None, "from_cache": True}

    # 2. Pas de cache valide — archiver l'ancien si présent (même clé, fichier disparu)
    replaced = None
    for existing in list(index["reports"]):
        if (
            existing.get("dedup_key") == dedup_key_str
            and existing.get("format") == fmt
        ):
            replaced = existing["filename"]
            old_path = _find_report_path(replaced)
            if old_path and old_path.exists():
                _archive_report(old_path, existing)
            index["reports"].remove(existing)

    # 3. Construire le filename + générer
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    slug = _slugify(title_str)
    ext_map = {"pdf": ".pdf", "csv": ".csv", "xlsx": ".xlsx", "excel": ".xlsx"}
    ext = ext_map.get(fmt, f".{fmt}")
    filename = f"rapport_{slug}_{timestamp}{ext}"
    output_path = REPORTS_DIR / filename
    output_path.parent.mkdir(parents=True, exist_ok=True)

    year = filters.get("year")
    template["renderer"](year=year, output_path=output_path, format=fmt, filters=filters)

    stat = output_path.stat()
    meta = {
        "filename": filename,
        "title": title_str,
        "description": description,
        "format": fmt,
        "generated_at": datetime.now().isoformat(),
        "filters": filters,
        "template_id": template_id,
        "dedup_key": dedup_key_str,
        "nb_operations": 0,
        "total_debit": 0,
        "total_credit": 0,
        "file_size": stat.st_size,
        "file_size_human": _format_size(stat.st_size),
        "year": year,
        "quarter": None,
        "month": filters.get("month"),
    }
    index["reports"].append(meta)
    _save_index(index)

    # 4. Enregistrement GED V2 (best-effort)
    try:
        from backend.services import ged_service
        ged_service.register_rapport(
            filename=filename,
            path=str(output_path),
            title=title_str,
            description=description,
            filters=filters,
            format_type=fmt,
            template_id=template_id,
            replaced_filename=replaced,
        )
    except Exception as e:
        logger.warning("Enregistrement GED rapport %s échoué: %s", filename, e)

    logger.info(
        "get_or_generate generated: %s (%s, format=%s, replaced=%s)",
        filename, dedup_key_str, fmt, replaced,
    )
    return {**meta, "replaced": replaced, "from_cache": False}


def generate_report(request: dict) -> dict:
    """Génère un rapport V2 avec index, déduplication, format EUR.

    Si `template_id` correspond à un template avec `renderer` custom (amortissements),
    délègue à `get_or_generate` qui dispatche sur le moteur dédié.
    """
    ensure_directories()

    fmt = request.get("format", "pdf")
    filters = request.get("filters", {})
    template_id = request.get("template_id")
    title = request.get("title")
    description = request.get("description")

    # Dispatch templates « custom renderer » (Prompt B3 — amortissements)
    if template_id:
        template = _get_template_by_id(template_id)
        if template and template.get("renderer"):
            return get_or_generate(
                template_id=template_id,
                filters=filters,
                format=fmt,
                title=title,
                description=description,
            )

    # Legacy path (BNC annuel / Ventilation charges / Récap social / freeform)
    if not title:
        title = _generate_title(filters, template_id)

    # Load operations
    year = filters.get("year")
    quarter = filters.get("quarter")
    month = filters.get("month")

    files = list_operation_files()
    if year:
        files = [f for f in files if f.get("year") == year]
    if month:
        files = [f for f in files if f.get("month") == month]
    if quarter:
        q_start = (quarter - 1) * 3 + 1
        files = [f for f in files if q_start <= (f.get("month") or 0) <= q_start + 2]

    all_ops: list[dict] = []
    for f in files:
        try:
            ops = load_operations(f["filename"])
            all_ops.extend(ops)
        except Exception:
            continue

    if not all_ops:
        # If no year filter, load all
        if not year:
            for f in list_operation_files():
                try:
                    ops = load_operations(f["filename"])
                    all_ops.extend(ops)
                except Exception:
                    continue

    # Éclater les ventilations AVANT les filtres pour que les sous-catégories soient filtrables
    exploded = _explode_ventilations(all_ops)

    # Apply filters
    filtered = _apply_filters(exploded, filters)

    if not filtered:
        raise ValueError("Aucune opération après application des filtres")

    # Deduplication check — archive old version instead of deleting
    replaced = None
    key = _dedup_key(filters, fmt)
    index = _load_index()
    for existing in index["reports"]:
        existing_key = _dedup_key(existing.get("filters", {}), existing.get("format", ""))
        if existing_key == key:
            replaced = existing["filename"]
            old_path = _find_report_path(replaced)
            if old_path and old_path.exists():
                _archive_report(old_path, existing)
            index["reports"] = [r for r in index["reports"] if r["filename"] != replaced]
            break

    # Generate file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    slug = _slugify(title)
    ext = {"pdf": ".pdf", "csv": ".csv", "excel": ".xlsx"}.get(fmt, ".pdf")
    filename = f"rapport_{slug}_{timestamp}{ext}"
    filepath = REPORTS_DIR / filename

    if fmt == "csv":
        _generate_csv_v2(filtered, filepath, title, filters)
    elif fmt == "pdf":
        _generate_pdf_v2(filtered, filepath, title, filters)
    elif fmt == "excel":
        _generate_excel_v2(filtered, filepath, title, filters)
    else:
        raise ValueError(f"Format non supporté: {fmt}")

    # Stats
    total_debit = sum(op.get("Débit", 0) for op in filtered)
    total_credit = sum(op.get("Crédit", 0) for op in filtered)
    stat = filepath.stat()

    # Add to index
    meta = {
        "filename": filename,
        "title": title,
        "description": description,
        "format": fmt,
        "generated_at": datetime.now().isoformat(),
        "filters": filters,
        "template_id": template_id,
        "nb_operations": len(filtered),
        "total_debit": round(total_debit, 2),
        "total_credit": round(total_credit, 2),
        "file_size": stat.st_size,
        "file_size_human": _format_size(stat.st_size),
        "year": year,
        "quarter": quarter,
        "month": month,
    }
    index["reports"].append(meta)
    _save_index(index)

    # Register in GED V2
    try:
        from backend.services import ged_service
        ged_service.register_rapport(
            filename=filename,
            path=str(filepath),
            title=title,
            description=description,
            filters=filters,
            format_type=fmt,
            template_id=template_id,
            replaced_filename=replaced,
        )
    except Exception:
        pass

    result = {**meta, "replaced": replaced}
    return result


def regenerate_report(filename: str) -> dict:
    """Régénère un rapport existant en gardant titre et description."""
    index = _load_index()
    existing = None
    for r in index["reports"]:
        if r["filename"] == filename:
            existing = r
            break
    if not existing:
        raise ValueError(f"Rapport non trouvé: {filename}")

    # Regenerate with same params
    request = {
        "format": existing["format"],
        "title": existing["title"],
        "description": existing.get("description"),
        "filters": existing.get("filters", {}),
        "template_id": existing.get("template_id"),
    }
    # Remove old entry to avoid dedup removing itself
    index["reports"] = [r for r in index["reports"] if r["filename"] != filename]
    _save_index(index)
    # Delete old file
    old_path = _find_report_path(filename)
    if old_path and old_path.exists():
        old_path.unlink()

    return generate_report(request)


# ─── Filters ───

def _explode_ventilations(operations: list[dict]) -> list[dict]:
    """Éclate les opérations ventilées en N sous-lignes (une par sous-ligne de ventilation).

    Chaque sous-ligne produite contient :
    - Date, Commentaire, Important : hérités du parent
    - Libellé : `<parent> [V{i+1}/{N}]`
    - Débit/Crédit : `vl.montant` projeté sur le côté d'origine (débit OU crédit)
    - Catégorie/Sous-catégorie : de la sous-ligne (pas du parent "Ventilé")
    - Justificatif : bool basé sur `vl.justificatif`
    - Lien justificatif : nom de fichier de la sous-ligne (si présent)

    Les ops non ventilées sont passées inchangées. Les totaux par catégorie sont
    ainsi correctement répartis entre les sous-catégories au lieu d'être agrégés
    sur la catégorie parente "Ventilé".
    """
    out: list[dict] = []
    for op in operations:
        vlines = op.get("ventilation") or []
        if not vlines:
            out.append(op)
            continue
        parent_libelle = op.get("Libellé", "") or ""
        parent_debit_positive = float(op.get("Débit", 0) or 0) > 0
        parent_credit_positive = float(op.get("Crédit", 0) or 0) > 0
        n = len(vlines)
        for i, vl in enumerate(vlines):
            montant = float(vl.get("montant", 0) or 0)
            vl_justif = (vl.get("justificatif") or "").strip()
            out.append({
                "Date": op.get("Date", ""),
                "Libellé": f"{parent_libelle} [V{i+1}/{n}]",
                "Catégorie": vl.get("categorie", "") or "",
                "Sous-catégorie": vl.get("sous_categorie", "") or "",
                "Débit": montant if parent_debit_positive else 0,
                "Crédit": montant if parent_credit_positive else 0,
                "Justificatif": bool(vl_justif),
                "Lien justificatif": f"traites/{vl_justif}" if vl_justif else "",
                "Commentaire": op.get("Commentaire", "") or "",
                "Important": op.get("Important", False),
            })
    return out


def _apply_filters(operations: list[dict], filters: dict) -> list[dict]:
    result = operations

    cats = filters.get("categories")
    if cats:
        include_uncategorized = "__non_categorise__" in cats
        real_cats = [c for c in cats if c != "__non_categorise__"]
        result = [
            op for op in result
            if op.get("Catégorie") in real_cats
            or (include_uncategorized and not op.get("Catégorie"))
            or (include_uncategorized and op.get("Catégorie") in ("", "Autres"))
        ]

    subcats = filters.get("subcategories")
    if subcats:
        result = [op for op in result if op.get("Sous-catégorie") in subcats]

    if filters.get("date_from"):
        result = [op for op in result if (op.get("Date") or "") >= filters["date_from"]]
    if filters.get("date_to"):
        result = [op for op in result if (op.get("Date") or "") <= filters["date_to"]]

    op_type = filters.get("type")
    if op_type == "debit":
        result = [op for op in result if op.get("Débit", 0) > 0]
    elif op_type == "credit":
        result = [op for op in result if op.get("Crédit", 0) > 0]

    source_filter = filters.get("source")
    if source_filter == "note_de_frais":
        result = [op for op in result if op.get("source") == "note_de_frais"]
    elif source_filter == "bancaire":
        result = [op for op in result if not op.get("source")]

    if filters.get("important_only"):
        result = [op for op in result if op.get("Important")]

    if filters.get("min_amount"):
        min_a = float(filters["min_amount"])
        result = [op for op in result if max(op.get("Débit", 0), op.get("Crédit", 0)) >= min_a]
    if filters.get("max_amount"):
        max_a = float(filters["max_amount"])
        result = [op for op in result if max(op.get("Débit", 0), op.get("Crédit", 0)) <= max_a]

    # Justificatif present / absent — utilisé par le template "compte_attente_sans_justif"
    # ainsi que par les rapports filtrant sur le scope justif. Aligné sur la règle d'unicité
    # frontend (cf. CLAUDE.md > Pipeline ↔ Justificatifs) : présence = `Lien justificatif`
    # non vide. Pour les ops ventilées éclatées, `_explode_ventilations` recopie déjà la
    # bonne valeur dans chaque sous-ligne.
    justif_present = filters.get("justificatif_present")
    if justif_present is True:
        result = [op for op in result if (op.get("Lien justificatif") or "").strip()]
    elif justif_present is False:
        result = [op for op in result if not (op.get("Lien justificatif") or "").strip()]

    return result


# ─── Ventilation par sous-catégorie ───
#
# Helper partagé par les 3 générateurs (PDF / CSV / Excel). Agrège les ops par
# `Catégorie → Sous-catégorie` avec compteurs + sommes débit/crédit. Utilisé
# pour la section « Ventilation par sous-catégorie » qui s'affiche quand
# au moins une opération a une sous-catégorie renseignée OU quand l'utilisateur
# a explicitement sélectionné des sous-catégories dans les filtres (cf.
# ReportFilters > Sous-catégories). Ordre stable : alphabétique.

def _aggregate_by_cat_subcat(operations: list[dict]) -> dict:
    """Retourne `{cat: {"subcats": {subcat: {debit, credit, count}}, "total_debit", "total_credit", "count"}}`."""
    out: dict[str, dict] = {}
    for op in operations:
        cat = (op.get("Catégorie") or "Non catégorisé").strip() or "Non catégorisé"
        subcat = (op.get("Sous-catégorie") or "").strip() or "(non précisée)"
        debit = float(op.get("Débit", 0) or 0)
        credit = float(op.get("Crédit", 0) or 0)
        cat_entry = out.setdefault(cat, {"subcats": {}, "total_debit": 0.0, "total_credit": 0.0, "count": 0})
        sub_entry = cat_entry["subcats"].setdefault(subcat, {"debit": 0.0, "credit": 0.0, "count": 0})
        sub_entry["debit"] += debit
        sub_entry["credit"] += credit
        sub_entry["count"] += 1
        cat_entry["total_debit"] += debit
        cat_entry["total_credit"] += credit
        cat_entry["count"] += 1
    return out


def _should_show_subcat_breakdown(operations: list[dict], filters: Optional[dict]) -> bool:
    """Active la section ventilation sous-catégorie ssi :
    - l'utilisateur a sélectionné des sous-catégories dans les filtres, OU
    - au moins une opération a une sous-catégorie renseignée.
    """
    f = filters or {}
    if f.get("subcategories"):
        return True
    return any((op.get("Sous-catégorie") or "").strip() for op in operations)


# ─── CSV Generation ───

def _generate_csv_v2(operations: list[dict], filepath: Path, title: str, filters: Optional[dict] = None):
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        f.write("Date;Libellé;Catégorie;Sous-catégorie;Débit;Crédit;Justificatif;Commentaire\n")
        for op in operations:
            date = op.get("Date", "")
            lib = str(op.get("Libellé", "")).replace(";", ",")
            cat = str(op.get("Catégorie", ""))
            scat = str(op.get("Sous-catégorie", ""))
            debit = f"{op.get('Débit', 0):.2f}".replace(".", ",") if op.get("Débit", 0) else ""
            credit = f"{op.get('Crédit', 0):.2f}".replace(".", ",") if op.get("Crédit", 0) else ""
            lien = op.get("Lien justificatif", "") or ""
            just_name = os.path.basename(lien) if lien else ""
            comment = str(op.get("Commentaire", "") or "").replace(";", ",")
            f.write(f"{date};{lib};{cat};{scat};{debit};{credit};{just_name};{comment}\n")

        total_d = sum(op.get("Débit", 0) for op in operations)
        total_c = sum(op.get("Crédit", 0) for op in operations)
        nb_just = sum(1 for op in operations if op.get("Lien justificatif"))
        f.write(f";TOTAUX;;;{total_d:.2f};{total_c:.2f};{nb_just}/{len(operations)};\n".replace(".", ","))

        # ── Ventilation par sous-catégorie ──
        if _should_show_subcat_breakdown(operations, filters):
            f.write("\n")
            f.write("Ventilation par sous-catégorie\n")
            f.write("Catégorie;Sous-catégorie;Nb ops;Total Débit;Total Crédit;Solde\n")
            agg = _aggregate_by_cat_subcat(operations)
            grand_d = 0.0
            grand_c = 0.0
            grand_n = 0
            for cat in sorted(agg.keys()):
                cat_entry = agg[cat]
                # Lignes sous-cat
                for subcat in sorted(cat_entry["subcats"].keys()):
                    s = cat_entry["subcats"][subcat]
                    solde = s["credit"] - s["debit"]
                    f.write(
                        f"{cat};{subcat};{s['count']};"
                        f"{s['debit']:.2f};{s['credit']:.2f};{solde:.2f}\n".replace(".", ",")
                    )
                # Sous-total catégorie
                cat_solde = cat_entry["total_credit"] - cat_entry["total_debit"]
                f.write(
                    f"{cat};Sous-total {cat};{cat_entry['count']};"
                    f"{cat_entry['total_debit']:.2f};{cat_entry['total_credit']:.2f};"
                    f"{cat_solde:.2f}\n".replace(".", ",")
                )
                grand_d += cat_entry["total_debit"]
                grand_c += cat_entry["total_credit"]
                grand_n += cat_entry["count"]
            grand_solde = grand_c - grand_d
            f.write(
                f";TOTAL GÉNÉRAL;{grand_n};"
                f"{grand_d:.2f};{grand_c:.2f};{grand_solde:.2f}\n".replace(".", ",")
            )


# ─── PDF Generation ───

def _generate_pdf_v2(operations: list[dict], filepath: Path, title: str, filters: dict):
    try:
        from reportlab.lib import colors
        from reportlab.lib.colors import HexColor
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import cm, mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    except ImportError:
        raise RuntimeError("reportlab non installé")

    PRIMARY = HexColor("#811971")

    doc = SimpleDocTemplate(
        str(filepath), pagesize=landscape(A4),
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=2 * cm, bottomMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("RTitle", parent=styles["Heading1"], fontSize=14,
                                  spaceAfter=4, textColor=PRIMARY)
    subtitle_style = ParagraphStyle("RSub", parent=styles["Normal"], fontSize=8,
                                     textColor=HexColor("#888888"), spaceAfter=8)
    small_style = ParagraphStyle("RSmall", parent=styles["Normal"], fontSize=7, leading=9)

    elements = []

    # Logo
    logo_path = ASSETS_DIR / "logo_lockup_light_400.png"
    if logo_path.exists():
        try:
            logo = Image(str(logo_path), width=120, height=40)
            logo.hAlign = "LEFT"
            elements.append(logo)
            elements.append(Spacer(1, 6))
        except Exception:
            pass  # Graceful fallback if logo can't be loaded

    # Header
    elements.append(Paragraph(title, title_style))
    filter_desc = _describe_filters(filters)
    elements.append(Paragraph(
        f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} | {len(operations)} opérations | {filter_desc}",
        subtitle_style
    ))
    elements.append(Spacer(1, 8))

    # Justificatif styles
    from reportlab.lib.enums import TA_CENTER
    just_yes_style = ParagraphStyle("RJustYes", parent=small_style, textColor=HexColor("#16a34a"),
                                     alignment=TA_CENTER)
    just_no_style = ParagraphStyle("RJustNo", parent=small_style, textColor=HexColor("#aaaaaa"),
                                    alignment=TA_CENTER)
    just_name_style = ParagraphStyle("RJustName", parent=small_style, fontSize=5.5, leading=7,
                                      textColor=HexColor("#666666"))

    # Comment style
    comment_style = ParagraphStyle("RComment", parent=small_style, fontSize=6, leading=7,
                                    textColor=HexColor("#555555"), fontName="Helvetica-Oblique")

    # Table
    headers = ["Date", "Libellé", "Catégorie", "Sous-cat.", "Débit", "Crédit", "Just.", "Commentaire"]
    table_data = [headers]

    for op in operations[:500]:
        debit = op.get("Débit", 0)
        credit = op.get("Crédit", 0)
        lien = op.get("Lien justificatif", "") or ""
        has_just = bool(lien)
        just_name = os.path.basename(lien) if lien else ""
        comment = str(op.get("Commentaire", "") or "")
        # Checkbox ☑ or ☐ + filename on second line
        if has_just:
            just_cell = Paragraph(f"☑ <font size=5>{just_name[:25]}</font>", just_yes_style)
        else:
            just_cell = Paragraph("☐", just_no_style)
        row = [
            str(op.get("Date", "")),
            Paragraph(str(op.get("Libellé", ""))[:80], small_style),
            str(op.get("Catégorie", "")),
            str(op.get("Sous-catégorie", "")),
            _format_eur(debit) if debit else "",
            _format_eur(credit) if credit else "",
            just_cell,
            Paragraph(comment[:40], comment_style) if comment else "",
        ]
        table_data.append(row)

    # Totals row
    total_d = sum(op.get("Débit", 0) for op in operations)
    total_c = sum(op.get("Crédit", 0) for op in operations)
    nb_just = sum(1 for op in operations if op.get("Lien justificatif"))
    table_data.append(["", "TOTAUX", "", "", _format_eur(total_d), _format_eur(total_c),
                        f"{nb_just}/{len(operations)}", ""])

    col_widths = [2 * cm, 6.5 * cm, 2.8 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm, 2.8 * cm, 2.7 * cm]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    n_rows = len(table_data)
    style_cmds = [
        # Header
        ("BACKGROUND", (0, 0), (-1, 0), PRIMARY),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        # Body
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("ALIGN", (4, 1), (5, -1), "RIGHT"),
        ("ALIGN", (6, 0), (6, 0), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#dddddd")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, HexColor("#f5f5f5")]),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        # Totals row
        ("BACKGROUND", (0, n_rows - 1), (-1, n_rows - 1), HexColor("#f0e8ef")),
        ("FONTNAME", (0, n_rows - 1), (-1, n_rows - 1), "Helvetica-Bold"),
        ("FONTSIZE", (0, n_rows - 1), (-1, n_rows - 1), 8),
        ("LINEABOVE", (0, n_rows - 1), (-1, n_rows - 1), 1.5, PRIMARY),
    ]
    table.setStyle(TableStyle(style_cmds))
    elements.append(table)

    if len(operations) > 500:
        elements.append(Spacer(1, 8))
        elements.append(Paragraph(
            f"Note : seules les 500 premières opérations sur {len(operations)} sont affichées.",
            subtitle_style
        ))

    # ── Section « Ventilation par sous-catégorie » ──
    # Affichée si l'utilisateur a sélectionné des sous-cat dans les filtres OU si
    # au moins une op a une sous-cat renseignée. Style miroir de la section
    # « Ventilation par catégorie » des exports comptables (header bleu foncé,
    # sous-total catégorie en bandeau, total général en bandeau noir).
    if _should_show_subcat_breakdown(operations, filters):
        elements.append(Spacer(1, 14))
        section_title_style = ParagraphStyle(
            "RSecTitle", parent=styles["Heading2"], fontSize=12,
            textColor=PRIMARY, spaceAfter=4,
        )
        elements.append(Paragraph("Ventilation par sous-catégorie", section_title_style))

        agg = _aggregate_by_cat_subcat(operations)
        sub_headers = ["Catégorie", "Sous-catégorie", "Nb ops", "Total Débit", "Total Crédit", "Solde"]
        sub_data: list[list] = [sub_headers]
        # Index lignes pour styler sous-totaux et total général
        sub_total_rows: list[int] = []
        grand_d = 0.0
        grand_c = 0.0
        grand_n = 0
        cur_row = 1
        for cat in sorted(agg.keys()):
            cat_entry = agg[cat]
            for subcat in sorted(cat_entry["subcats"].keys()):
                s = cat_entry["subcats"][subcat]
                solde = s["credit"] - s["debit"]
                sub_data.append([
                    cat,
                    subcat,
                    str(s["count"]),
                    _format_eur(s["debit"]) if s["debit"] else "—",
                    _format_eur(s["credit"]) if s["credit"] else "—",
                    _format_eur(solde) if solde else "—",
                ])
                cur_row += 1
            # Sous-total catégorie
            cat_solde = cat_entry["total_credit"] - cat_entry["total_debit"]
            sub_data.append([
                "",
                f"Sous-total {cat}",
                str(cat_entry["count"]),
                _format_eur(cat_entry["total_debit"]),
                _format_eur(cat_entry["total_credit"]),
                _format_eur(cat_solde),
            ])
            sub_total_rows.append(cur_row)
            cur_row += 1
            grand_d += cat_entry["total_debit"]
            grand_c += cat_entry["total_credit"]
            grand_n += cat_entry["count"]
        # Total général
        grand_solde = grand_c - grand_d
        sub_data.append([
            "TOTAL GÉNÉRAL",
            "",
            str(grand_n),
            _format_eur(grand_d),
            _format_eur(grand_c),
            _format_eur(grand_solde),
        ])
        grand_total_row = cur_row

        sub_col_widths = [4.5 * cm, 6 * cm, 2 * cm, 3 * cm, 3 * cm, 3 * cm]
        sub_table = Table(sub_data, colWidths=sub_col_widths, repeatRows=1)
        sub_style: list = [
            # Header (bleu foncé pour distinguer de la table principale violette)
            ("BACKGROUND", (0, 0), (-1, 0), HexColor("#1f4e79")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            # Body
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.4, HexColor("#dddddd")),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]
        # Sous-totaux en bandeau bleu clair
        for r in sub_total_rows:
            sub_style.append(("BACKGROUND", (0, r), (-1, r), HexColor("#dbe9f4")))
            sub_style.append(("FONTNAME", (0, r), (-1, r), "Helvetica-Bold"))
            sub_style.append(("LINEABOVE", (0, r), (-1, r), 0.6, HexColor("#1f4e79")))
        # Total général en bandeau sombre
        sub_style.append(("BACKGROUND", (0, grand_total_row), (-1, grand_total_row), HexColor("#1a1a1a")))
        sub_style.append(("TEXTCOLOR", (0, grand_total_row), (-1, grand_total_row), colors.white))
        sub_style.append(("FONTNAME", (0, grand_total_row), (-1, grand_total_row), "Helvetica-Bold"))
        sub_style.append(("FONTSIZE", (0, grand_total_row), (-1, grand_total_row), 9))
        sub_table.setStyle(TableStyle(sub_style))
        elements.append(sub_table)

    doc.build(elements)


def _describe_filters(filters: dict) -> str:
    parts = []
    if filters.get("categories"):
        parts.append(", ".join(filters["categories"]))
    if filters.get("type") == "debit":
        parts.append("Dépenses")
    elif filters.get("type") == "credit":
        parts.append("Recettes")
    return " | ".join(parts) if parts else "Tous"


# ─── Excel Generation ───

def _generate_excel_v2(operations: list[dict], filepath: Path, title: str, filters: Optional[dict] = None):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    except ImportError:
        raise RuntimeError("openpyxl non installé")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Opérations"

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="811971", end_color="811971", fill_type="solid")
    total_fill = PatternFill(start_color="F0E8EF", end_color="F0E8EF", fill_type="solid")
    border = Border(
        left=Side(style="thin", color="DDDDDD"), right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"), bottom=Side(style="thin", color="DDDDDD"),
    )

    columns = ["Date", "Libellé", "Catégorie", "Sous-catégorie", "Débit", "Crédit"]

    for ci, name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=ci, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for ri, op in enumerate(operations, 2):
        for ci, col in enumerate(columns, 1):
            val = op.get(col, "")
            if col in ("Débit", "Crédit"):
                val = val if val else 0
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = border
            if col in ("Débit", "Crédit"):
                cell.number_format = '#,##0.00 €'
                cell.alignment = Alignment(horizontal="right")

    # Totals row
    n = len(operations) + 2
    ws.cell(row=n, column=2, value="TOTAUX").font = Font(bold=True)
    for ci in (5, 6):
        col_letter = openpyxl.utils.get_column_letter(ci)
        cell = ws.cell(row=n, column=ci, value=f"=SUM({col_letter}2:{col_letter}{n - 1})")
        cell.font = Font(bold=True)
        cell.fill = total_fill
        cell.number_format = '#,##0.00 €'
        cell.border = border

    widths = [12, 40, 18, 18, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # Sheet 2: Category summary (if > 10 ops)
    if len(operations) > 10:
        ws2 = wb.create_sheet("Résumé")
        cat_stats: dict[str, dict] = {}
        for op in operations:
            cat = op.get("Catégorie", "Non catégorisé") or "Non catégorisé"
            if cat not in cat_stats:
                cat_stats[cat] = {"debit": 0, "credit": 0, "count": 0}
            cat_stats[cat]["debit"] += op.get("Débit", 0)
            cat_stats[cat]["credit"] += op.get("Crédit", 0)
            cat_stats[cat]["count"] += 1

        h2 = ["Catégorie", "Nb ops", "Total Débit", "Total Crédit", "Solde"]
        for ci, name in enumerate(h2, 1):
            cell = ws2.cell(row=1, column=ci, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        for ri, (cat, s) in enumerate(sorted(cat_stats.items()), 2):
            vals = [cat, s["count"], s["debit"], s["credit"], s["credit"] - s["debit"]]
            for ci, v in enumerate(vals, 1):
                cell = ws2.cell(row=ri, column=ci, value=v)
                cell.border = border
                if ci >= 3:
                    cell.number_format = '#,##0.00 €'

        for i, w in enumerate([25, 10, 15, 15, 15], 1):
            ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws2.freeze_panes = "A2"

    # Sheet 3: Ventilation par sous-catégorie (cat → sous-cat avec sous-totaux)
    if _should_show_subcat_breakdown(operations, filters):
        ws3 = wb.create_sheet("Ventilation sous-cat")
        ws3.append(["Catégorie", "Sous-catégorie", "Nb ops", "Total Débit", "Total Crédit", "Solde"])
        # Header style — bleu foncé pour distinguer de l'onglet "Résumé" (violet)
        header_blue_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        for cell in ws3[1]:
            cell.font = header_font
            cell.fill = header_blue_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        agg = _aggregate_by_cat_subcat(operations)
        sub_total_fill = PatternFill(start_color="DBE9F4", end_color="DBE9F4", fill_type="solid")
        grand_total_fill = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
        grand_total_font = Font(bold=True, color="FFFFFF")
        grand_d = 0.0
        grand_c = 0.0
        grand_n = 0
        for cat in sorted(agg.keys()):
            cat_entry = agg[cat]
            for subcat in sorted(cat_entry["subcats"].keys()):
                s = cat_entry["subcats"][subcat]
                solde = s["credit"] - s["debit"]
                ws3.append([cat, subcat, s["count"], s["debit"], s["credit"], solde])
                # Format colonnes monétaires
                row_idx = ws3.max_row
                for ci in (4, 5, 6):
                    c = ws3.cell(row=row_idx, column=ci)
                    c.number_format = '#,##0.00 €'
                    c.alignment = Alignment(horizontal="right")
                    c.border = border
                for ci in (1, 2, 3):
                    ws3.cell(row=row_idx, column=ci).border = border
            # Sous-total catégorie
            cat_solde = cat_entry["total_credit"] - cat_entry["total_debit"]
            ws3.append([
                "",
                f"Sous-total {cat}",
                cat_entry["count"],
                cat_entry["total_debit"],
                cat_entry["total_credit"],
                cat_solde,
            ])
            sub_row = ws3.max_row
            for ci in range(1, 7):
                c = ws3.cell(row=sub_row, column=ci)
                c.fill = sub_total_fill
                c.font = Font(bold=True)
                c.border = border
                if ci >= 4:
                    c.number_format = '#,##0.00 €'
                    c.alignment = Alignment(horizontal="right")
            grand_d += cat_entry["total_debit"]
            grand_c += cat_entry["total_credit"]
            grand_n += cat_entry["count"]
        # Total général en bandeau noir
        grand_solde = grand_c - grand_d
        ws3.append(["TOTAL GÉNÉRAL", "", grand_n, grand_d, grand_c, grand_solde])
        grand_row = ws3.max_row
        for ci in range(1, 7):
            c = ws3.cell(row=grand_row, column=ci)
            c.fill = grand_total_fill
            c.font = grand_total_font
            c.border = border
            if ci >= 4:
                c.number_format = '#,##0.00 €'
                c.alignment = Alignment(horizontal="right")
        for i, w in enumerate([22, 28, 10, 16, 16, 16], 1):
            ws3.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws3.freeze_panes = "A2"

    wb.save(str(filepath))


# ─── Helpers ───

def _format_size(size_bytes: int) -> str:
    if size_bytes < 1024:
        return f"{size_bytes} o"
    elif size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.1f} Ko"
    else:
        return f"{size_bytes / (1024 * 1024):.1f} Mo"


# ─── Favorites ───

def toggle_favorite(filename: str) -> Optional[dict]:
    index = _load_index()
    for r in index["reports"]:
        if r["filename"] == filename:
            r["favorite"] = not r.get("favorite", False)
            _save_index(index)
            return r
    return None


# ─── Report Tree (triple vue) ───

def get_report_tree() -> dict:
    """Construit un arbre triple vue : by_year, by_category, by_format."""
    reports = get_all_reports()

    # By year
    by_year_map: dict[int, list[dict]] = {}
    no_year: list[dict] = []
    for r in reports:
        y = r.get("year")
        if y:
            by_year_map.setdefault(y, []).append(r)
        else:
            no_year.append(r)

    by_year = []
    for y in sorted(by_year_map.keys(), reverse=True):
        children = []
        # Group by quarter/month
        by_month: dict[int, int] = {}
        for r in by_year_map[y]:
            m = r.get("month") or 0
            by_month[m] = by_month.get(m, 0) + 1
        for m in sorted(by_month.keys()):
            label = MOIS_FR[m - 1].capitalize() if 1 <= m <= 12 else "Général"
            children.append({"id": f"year-{y}-{m}", "label": label, "count": by_month[m], "children": []})
        by_year.append({"id": f"year-{y}", "label": str(y), "count": len(by_year_map[y]), "children": children, "icon": "Calendar"})
    if no_year:
        by_year.append({"id": "year-none", "label": "Non daté", "count": len(no_year), "children": []})

    # By category
    by_cat_map: dict[str, int] = {}
    no_cat = 0
    for r in reports:
        cats = (r.get("filters") or {}).get("categories")
        if cats:
            for c in cats:
                by_cat_map[c] = by_cat_map.get(c, 0) + 1
        else:
            no_cat += 1

    by_category = []
    for cat in sorted(by_cat_map.keys()):
        by_category.append({"id": f"cat-{cat}", "label": cat, "count": by_cat_map[cat], "children": [], "icon": "Tag"})
    if no_cat:
        by_category.append({"id": "cat-none", "label": "Toutes catégories", "count": no_cat, "children": []})

    # By format
    by_fmt_map: dict[str, int] = {}
    for r in reports:
        fmt = r.get("format", "pdf")
        by_fmt_map[fmt] = by_fmt_map.get(fmt, 0) + 1

    fmt_icons = {"pdf": "FileText", "csv": "Sheet", "excel": "Table2"}
    by_format = []
    for fmt in ["pdf", "csv", "excel"]:
        if fmt in by_fmt_map:
            by_format.append({"id": f"fmt-{fmt}", "label": fmt.upper(), "count": by_fmt_map[fmt], "children": [], "icon": fmt_icons.get(fmt, "FileText")})

    return {"by_year": by_year, "by_category": by_category, "by_format": by_format}


# ─── Pending Reports (rappels) ───

def get_pending_reports(year: int) -> list[dict]:
    """Retourne les rapports mensuels manquants pour les mois passés."""
    now = datetime.now()
    current_month = now.month if year == now.year else (12 if year < now.year else 0)

    reports = get_all_reports()
    generated_months: set[int] = set()
    for r in reports:
        if r.get("year") == year and r.get("month"):
            generated_months.add(r["month"])

    pending = []
    for m in range(1, current_month + 1):
        if m not in generated_months:
            label = MOIS_FR[m - 1].capitalize()
            pending.append({
                "type": "mensuel",
                "period": f"{label} {year}",
                "message": f"Rapport {label} {year} non généré",
                "year": year,
                "month": m,
                "quarter": None,
            })

    # Check quarterly
    generated_quarters: set[int] = set()
    for r in reports:
        if r.get("year") == year and r.get("quarter"):
            generated_quarters.add(r["quarter"])

    current_q = (current_month - 1) // 3 + 1
    for q in range(1, current_q + 1):
        q_end_month = q * 3
        if q_end_month <= current_month and q not in generated_quarters:
            pending.append({
                "type": "trimestriel",
                "period": f"T{q} {year}",
                "message": f"Rapport trimestriel T{q} {year} non généré",
                "year": year,
                "month": None,
                "quarter": q,
            })

    return pending


# ─── Compare Reports ───

def compare_reports(filename_a: str, filename_b: str) -> dict:
    """Compare deux rapports et retourne les deltas."""
    index = _load_index()
    report_a = None
    report_b = None
    for r in index["reports"]:
        if r["filename"] == filename_a:
            report_a = r
        if r["filename"] == filename_b:
            report_b = r

    if not report_a or not report_b:
        raise ValueError("Un ou plusieurs rapports introuvables")

    delta_debit = report_a["total_debit"] - report_b["total_debit"]
    delta_credit = report_a["total_credit"] - report_b["total_credit"]
    delta_ops = report_a["nb_operations"] - report_b["nb_operations"]
    delta_debit_pct = (delta_debit / report_b["total_debit"] * 100) if report_b["total_debit"] else 0
    delta_credit_pct = (delta_credit / report_b["total_credit"] * 100) if report_b["total_credit"] else 0

    return {
        "report_a": report_a,
        "report_b": report_b,
        "delta_debit": round(delta_debit, 2),
        "delta_credit": round(delta_credit, 2),
        "delta_ops": delta_ops,
        "delta_debit_pct": round(delta_debit_pct, 1),
        "delta_credit_pct": round(delta_credit_pct, 1),
    }


# Legacy compatibility
def list_report_files() -> list[dict]:
    """Legacy: liste tous les rapports (compatible avec l'ancien code)."""
    return get_all_reports()
